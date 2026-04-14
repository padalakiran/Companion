# ── features/todo.py ──────────────────────────────────────────────────────────
# Patch notes:
# - Custom thin accent scrollbar
# - Mouse wheel scroll
# - Aligned priority icons and +Add button with entry height
# - Filter tab highlight on active
# - Placeholder clears on focus

import tkinter as tk
from tkinter import font as tkfont
from datetime import date
import os
import openpyxl
import config

import theme as _theme_mod
def _c(k): return _theme_mod.get(k)

# Persists across build() calls so placeholder only shows the very first time ever
_todo_placeholder_shown = [False]

def _ensure_wb():
    os.makedirs(config.DATA_DIR, exist_ok=True)
    if not os.path.exists(config.TODO_DATA):
        wb = openpyxl.Workbook(); ws = wb.active
        ws.title = "Todos"
        ws.append(["id","task","priority","due_date","done","created"])
        wb.save(config.TODO_DATA)

def load_todos():
    _ensure_wb()
    todos = []
    try:
        wb = openpyxl.load_workbook(config.TODO_DATA)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row[0]: continue
            todos.append({"id":row[0],"task":row[1] or "",
                          "priority":row[2] or "Normal",
                          "due_date":str(row[3]) if row[3] else "",
                          "done":bool(row[4]),
                          "created":str(row[5]) if row[5] else ""})
    except Exception as e:
        print(f"[todo] load: {e}")
    return todos

def save_todos(todos):
    _ensure_wb()
    try:
        wb = openpyxl.Workbook(); ws = wb.active
        ws.title = "Todos"
        ws.append(["id","task","priority","due_date","done","created"])
        for t in todos:
            ws.append([t["id"],t["task"],t["priority"],
                       t["due_date"],t["done"],t["created"]])
        wb.save(config.TODO_DATA)
    except Exception as e:
        print(f"[todo] save: {e}")

def _new_id(todos):
    return max((t["id"] for t in todos), default=0) + 1

def _custom_scrollbar(master, canvas):
    """4px accent scrollbar drawn on a Canvas."""
    SB_W = 4
    sb = tk.Canvas(master, width=SB_W, bg=_c("BG"), bd=0,
                   highlightthickness=0, cursor="arrow")
    sb.pack(side="right", fill="y", padx=(1,0))
    thumb = sb.create_rectangle(0, 0, SB_W, 40, fill=_c("ACCENT"), outline="")

    def update(*_):
        top, bot = canvas.yview()
        h = max(sb.winfo_height(), 1)
        sb.coords(thumb, 0, top*h, SB_W, max(bot*h, top*h+12))

    def drag(e):
        top, bot = canvas.yview()
        span = bot - top
        frac = e.y / max(sb.winfo_height(), 1)
        canvas.yview_moveto(max(0, frac - span/2))
        update()

    sb.bind("<ButtonPress-1>", drag)
    sb.bind("<B1-Motion>",     drag)
    canvas.bind("<<TodoScroll>>", update)
    return update

def build(parent: tk.Frame, root: tk.Tk):
    ref         = {"data": load_todos()}
    filter_var  = tk.StringVar(value="all")
    filter_btns = {}

    fh   = tkfont.Font(family="Segoe UI", size=11, weight="bold")
    fb   = tkfont.Font(family="Segoe UI", size=10)
    fs   = tkfont.Font(family="Segoe UI", size=9)
    fe   = tkfont.Font(family="Segoe UI", size=11)
    fbtn = tkfont.Font(family="Segoe UI", size=10, weight="bold")
    fpri = tkfont.Font(family="Segoe UI", size=9,  weight="bold")

    # ── Add bar ───────────────────────────────────────────────────────────────
    add_row = tk.Frame(parent, bg=_c("BG"))
    add_row.pack(fill="x", pady=(0,10))

    pri_var  = tk.StringVar(value="Normal")

    # Text entry — no textvariable so insert/delete don't conflict with StringVar
    entry_wrap = tk.Frame(add_row, bg=_c("BORDER"))
    entry_wrap.pack(side="left", fill="x", expand=True, padx=(0,6))
    entry = tk.Entry(entry_wrap,
                     font=fe, bg=_c("CARD"), fg=_c("SUB"),
                     insertbackground=_c("ACCENT"), relief="flat", bd=0)
    entry.pack(fill="x", padx=1, pady=1, ipady=8)

    # Show placeholder only on very first ever open (persists across navigation)
    _PLACEHOLDER = "Add a task…"

    if not _todo_placeholder_shown[0]:
        entry.insert(0, _PLACEHOLDER)   # show hint first time only
        entry.config(fg=_c("SUB"))

    def fi(e=None):
        if not _todo_placeholder_shown[0]:
            entry.delete(0, "end")
            entry.config(fg=_c("TEXT"))
            _todo_placeholder_shown[0] = True

    entry.bind("<FocusIn>",  fi)
    entry.bind("<Button-1>", fi)

    # Priority buttons — same ipady as entry so heights match
    pri_frame = tk.Frame(add_row, bg=_c("BG"))
    pri_frame.pack(side="left", padx=(0,6))

    pri_map   = {"!": "Low", "!!": "Normal", "!!!": "High"}
    pri_colors= {"!": _c("GREEN"), "!!": _c("YELLOW"), "!!!": _c("RED")}
    pri_refs  = {}

    def set_priority(val, lbl):
        pri_var.set(val)
        for l2, btn in pri_refs.items():
            is_sel = (pri_map[l2] == val)
            btn.config(bg=pri_colors[l2] if is_sel else _c("BORDER"),
                       fg=_c("BG") if is_sel else pri_colors[l2])

    for lbl in ("!", "!!", "!!!"):
        val = pri_map[lbl]
        btn = tk.Button(pri_frame, text=lbl, font=fpri,
                        bg=_c("BORDER"), fg=pri_colors[lbl],
                        activebackground=pri_colors[lbl], activeforeground=_c("BG"),
                        relief="flat", cursor="hand2", bd=0, width=3,
                        command=lambda v=val, l=lbl: set_priority(v, l))
        btn.pack(side="left", padx=1, ipady=8)
        pri_refs[lbl] = btn

    def add_task(*_):
        t = entry.get().strip()
        if not t:
            entry_wrap.config(bg=_c("RED"))
            parent.after(800, lambda: entry_wrap.config(bg=_c("BORDER")))
            return
        ref["data"].append({"id":_new_id(ref["data"]),"task":t,
                             "priority":pri_var.get(),"due_date":"",
                             "done":False,"created":str(date.today())})
        save_todos(ref["data"])
        entry.delete(0, "end")        # leave empty after submit — no placeholder
        set_priority("Normal", "!!")
        refresh()

    tk.Button(add_row, text="＋ Add", font=fbtn,
              bg=_c("ACCENT"), fg=_c("BG"),
              activebackground=_c("TEXT"), activeforeground=_c("BG"),
              relief="flat", cursor="hand2", bd=0,
              command=add_task).pack(side="left", ipady=8, ipadx=12)
    entry.bind("<Return>", add_task)

    # ── Filter tabs ───────────────────────────────────────────────────────────
    tab_row = tk.Frame(parent, bg=_c("BG"))
    tab_row.pack(fill="x", pady=(0,10))

    def set_filter(val):
        filter_var.set(val)
        for v2, btn in filter_btns.items():
            if v2 == val:
                btn.config(bg=_c("ACCENT"), fg=_c("BG"))
            else:
                btn.config(bg=_c("BORDER"), fg=_c("SUB"))
        refresh()

    for lbl, val in [("All","all"),("Pending","pending"),("Done","done")]:
        is_active = (val == "all")
        btn = tk.Button(tab_row, text=lbl, font=fs,
                        bg=_c("ACCENT") if is_active else _c("BORDER"),
                        fg=_c("BG")     if is_active else _c("SUB"),
                        activebackground=_c("ACCENT"), activeforeground=_c("BG"),
                        relief="flat", cursor="hand2", bd=0,
                        command=lambda v=val: set_filter(v))
        btn.pack(side="left", padx=(0,4), ipady=5, ipadx=14)
        filter_btns[val] = btn

    # ── List ──────────────────────────────────────────────────────────────────
    wrap = tk.Frame(parent, bg=_c("BG"))
    wrap.pack(fill="both", expand=True)

    # Static empty-state label — sits OUTSIDE the scrollable canvas
    # so it never scrolls with mouse movement
    empty_lbl = tk.Label(wrap, text="", font=fs, bg=_c("BG"), fg=_c("SUB"))
    empty_lbl.place(relx=0.5, rely=0.45, anchor="center")
    empty_lbl.lower()   # hidden by default behind canvas

    cv = tk.Canvas(wrap, bg=_c("BG"), bd=0, highlightthickness=0)
    cv.pack(side="left", fill="both", expand=True)
    lf = tk.Frame(cv, bg=_c("BG"))
    cwin = cv.create_window((0,0), window=lf, anchor="nw")

    sb_widget = _custom_scrollbar(wrap, cv)

    def _on_lf_configure(e):
        bb = cv.bbox("all")
        if bb:
            cv.configure(scrollregion=(0, 0, bb[2], bb[3]))
        cv.event_generate("<<TodoScroll>>")
    lf.bind("<Configure>", _on_lf_configure)
    cv.bind("<Configure>", lambda e: cv.itemconfig(cwin, width=e.width))
    def _scroll_if_needed(e):
        top, bot = cv.yview()
        if top == 0.0 and bot == 1.0: return
        cv.yview_scroll(-1*(e.delta//120), "units")
        cv.event_generate("<<TodoScroll>>")
    cv.bind("<MouseWheel>", _scroll_if_needed)
    lf.bind("<MouseWheel>", _scroll_if_needed)

    pcol = {"High":_c("RED"), "Normal":_c("ACCENT"), "Low":_c("GREEN")}

    def refresh():
        cv.yview_moveto(0)   # always start at top
        for w in lf.winfo_children(): w.destroy()
        filt  = filter_var.get()
        shown = [t for t in ref["data"]
                 if filt=="all"
                 or(filt=="done"    and t["done"])
                 or(filt=="pending" and not t["done"])]

        msgs = {"all":"No tasks yet — add one above!",
                "pending":"All done! Nothing pending.",
                "done":"No completed tasks yet."}
        if not shown:
            # Show static label, hide canvas (scrollbar disappears with it)
            empty_lbl.config(text=msgs.get(filt,""))
            empty_lbl.lift()
            cv.pack_forget()
            return
        # Tasks exist — hide empty label, show canvas
        empty_lbl.config(text="")
        empty_lbl.lower()
        if not cv.winfo_ismapped():
            cv.pack(side="left", fill="both", expand=True)

        for todo in shown:
            row_b = tk.Frame(lf, bg=_c("BORDER"))
            row_b.pack(fill="x", pady=2)
            row = tk.Frame(row_b, bg=_c("CARD"), height=44)
            row.pack(fill="x", padx=1, pady=1)
            row.pack_propagate(False)

            # Priority dot
            dc = tk.Canvas(row, width=8, height=8, bg=_c("CARD"),
                           bd=0, highlightthickness=0)
            dc.place(x=10, y=18)
            dc.create_oval(0,0,8,8, fill=pcol.get(todo["priority"],_c("ACCENT")), outline="")

            # Task text
            fg = _c("SUB") if todo["done"] else _c("TEXT")
            tl = tk.Label(row, text=todo["task"], font=fb,
                          bg=_c("CARD"), fg=fg, anchor="w", cursor="hand2")
            tl.place(x=26, y=12)

            # Done toggle
            done_lbl = tk.Label(row, text="✓" if todo["done"] else "○",
                                font=fh, bg=_c("CARD"),
                                fg=_c("GREEN") if todo["done"] else _c("BORDER"),
                                cursor="hand2")
            done_lbl.place(relx=1.0, x=-58, y=10)

            # Delete
            tk.Button(row, text="✕", font=fs, bg=_c("CARD"), fg=_c("RED"),
                      activebackground=_c("RED"), activeforeground=_c("BG"),
                      relief="flat", cursor="hand2", bd=0,
                      command=lambda tid=todo["id"]: delete(tid)
                      ).place(relx=1.0, x=-28, y=10)

            def toggle(tid=todo["id"]):
                for t in ref["data"]:
                    if t["id"] == tid:
                        t["done"] = not t["done"]; break
                save_todos(ref["data"]); refresh()

            for w in (row, tl, done_lbl):
                w.bind("<Button-1>", lambda e, fn=toggle: fn())
                w.bind("<MouseWheel>", _scroll_if_needed)

        cv.update_idletasks()
        bb = cv.bbox("all")
        if bb: cv.configure(scrollregion=(0, 0, bb[2], bb[3]))
        cv.event_generate("<<TodoScroll>>")

    def delete(tid):
        ref["data"] = [t for t in ref["data"] if t["id"] != tid]
        save_todos(ref["data"]); refresh()

    refresh()

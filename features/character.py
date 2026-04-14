# ── features/character.py ─────────────────────────────────────────────────────
# Phase 11: Character selector — spritesheet-based
#
# HOW TO ADD A CHARACTER:
# ───────────────────────
#   Drop a single PNG spritesheet into the  characters/  folder.
#
#   Required spritesheet layout:
#     • 5 rows × 4 columns × 150 px per cell  =  600 × 750 px total
#     • Row 0 — walk right  (4 frames, left→right motion)
#     • Row 1 — walk left   (4 frames, right→left motion)
#     • Row 2 — idle        (4 frames, gentle sway/bob)
#     • Row 3 — play        (4 frames, excited / click animation)
#     • Row 4 — stopped     (4 frames, standing still)
#
#   Cell co-ordinate
#       x1 = c * 150
#       y1 = r * 150
#       x2 = (c + 1) * 150
#       y2 = (r + 1) * 150
#   Cell size:    150 × 150 px
#   Full sheet:   600 × 750 px
#   Background:   Transparent (RGBA) recommended
#   Format:       PNG only
#   Naming:       Anything you like — shown as the character name in the gallery
# ─────────────────────────────────────────────────────────────────────────────

import tkinter as tk
from tkinter import font as tkfont
import os
import config
import theme as _theme

import theme as _theme_mod
def _c(k): return _theme_mod.get(k)

CHARACTERS_DIR = os.path.join(config.BASE_DIR, "characters")


# ── Helpers ───────────────────────────────────────────────────────────────────

def ensure_dir():
    os.makedirs(CHARACTERS_DIR, exist_ok=True)

def list_characters() -> list[dict]:
    ensure_dir()
    active = get_active_character()
    chars  = []

    # Built-in default
    chars.append({
        "name":      "Blue Cat (default)",
        "path":      config.SPRITESHEET,
        "is_active": active == config.SPRITESHEET,
    })

    for fname in sorted(os.listdir(CHARACTERS_DIR)):
        if not fname.lower().endswith(".png"):
            continue
        path = os.path.join(CHARACTERS_DIR, fname)
        name = os.path.splitext(fname)[0].replace("_"," ").title()
        chars.append({
            "name":      name,
            "path":      path,
            "is_active": os.path.abspath(path) == os.path.abspath(active),
        })
    return chars

def get_active_character() -> str:
    try:
        import openpyxl
        if os.path.exists(config.USER_DATA):
            wb = openpyxl.load_workbook(config.USER_DATA)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if row[0] == "active_character" and row[1]:
                    p = str(row[1])
                    if os.path.exists(p):
                        return p
    except Exception:
        pass
    return config.SPRITESHEET

def set_active_character(path: str):
    import openpyxl
    os.makedirs(config.DATA_DIR, exist_ok=True)
    try:
        if os.path.exists(config.USER_DATA):
            wb = openpyxl.load_workbook(config.USER_DATA)
            ws = wb.active
            for row in ws.iter_rows():
                if row[0].value == "active_character":
                    row[1].value = path
                    wb.save(config.USER_DATA); return
            ws.append(["active_character", path])
            wb.save(config.USER_DATA)
        else:
            wb = openpyxl.Workbook(); ws = wb.active
            ws.title = "User"
            ws.append(["key","value"])
            ws.append(["active_character", path])
            wb.save(config.USER_DATA)
    except Exception as e:
        print(f"[character] save: {e}")

def load_thumbnail(path: str, size: int = 100):
    """Crop the stopped frame (row 4, col 0) from spritesheet as thumbnail."""
    try:
        from PIL import Image, ImageTk
        sheet = Image.open(path).convert("RGBA")
        S     = config.SPRITE_SIZE
        # Use stopped frame (row 4) for thumbnail
        row   = 4
        frame = sheet.crop((0, row*S, S, (row+1)*S))
        frame = frame.resize((size, size), Image.LANCZOS)
        bg    = Image.new("RGBA", (size, size), (37,38,64,255))
        bg.paste(frame, (0,0), frame)
        return ImageTk.PhotoImage(bg)
    except Exception as e:
        print(f"[character] thumbnail: {e}")
        return None

# ── Main view ─────────────────────────────────────────────────────────────────

def build(parent: tk.Frame, root: tk.Tk, on_character_change=None):
    import theme; theme.refresh(globals())

    fh   = tkfont.Font(family="Segoe UI", size=12, weight="bold")
    fb   = tkfont.Font(family="Segoe UI", size=10)
    fs   = tkfont.Font(family="Segoe UI", size=9)
    fbn  = tkfont.Font(family="Segoe UI", size=10, weight="bold")
    fmono= tkfont.Font(family="Consolas",  size=8)

    tk.Label(parent, text="Choose your companion",
             font=fh, bg=_c("BG"), fg=_c("ACCENT")).pack(anchor="w", pady=(0,6))

    feedback = tk.Label(parent, text="", font=fs, bg=_c("BG"), fg=_c("GREEN"))
    feedback.pack(anchor="w")

    def show_msg(msg, color=_c("GREEN")):
        feedback.config(text=msg, fg=color)
        parent.after(3000, lambda: feedback.config(text=""))

    # ── Spritesheet spec card ─────────────────────────────────────────────────
    spec_b = tk.Frame(parent, bg=_c("BORDER"))
    spec_b.pack(fill="x", pady=(4,10))
    spec   = tk.Frame(spec_b, bg=_c("CARD"))
    spec.pack(fill="x", padx=1, pady=1)

    tk.Label(spec, text="➕  Add your own character",
             font=fbn, bg=_c("CARD"), fg=_c("YELLOW")).pack(anchor="w", padx=12, pady=(8,4))

    spec_text = (
        "Drop a PNG spritesheet into:  characters/  folder\n"
        "\n"
        "Required layout:\n"
        "  • 5 rows  ×  4 columns  ×  150 px per cell\n"
        "  • Total sheet size:  600 × 750 px\n"
        "\n"
        "  Row 0 — walk right   (4 frames)\n"
        "  Row 1 — walk left    (4 frames)\n"
        "  Row 2 — idle         (4 frames)\n"
        "  Row 3 — play / click (4 frames)\n"
        "  Row 4 — stopped      (4 frames)\n"
        "\n"
        "  Cell:   150 × 150 px  |  Format: PNG  |  BG: transparent"
    )
    tk.Label(spec, text=spec_text, font=fmono,
             bg=_c("CARD"), fg=_c("SUB"), justify="left", anchor="w"
             ).pack(anchor="w", padx=16, pady=(0,10))

    # ── Scrollable gallery ────────────────────────────────────────────────────
    outer = tk.Frame(parent, bg=_c("BG"))
    outer.pack(fill="both", expand=True)

    cv  = tk.Canvas(outer, bg=_c("BG"), bd=0, highlightthickness=0)
    cv.pack(side="left", fill="both", expand=True)
    SB_W = 4
    sb   = tk.Canvas(outer, width=SB_W, bg=_c("BG"), bd=0, highlightthickness=0)
    sb.pack(side="right", fill="y")
    tr   = sb.create_rectangle(0,0,SB_W,30, fill=_c("ACCENT"), outline="")

    gallery = tk.Frame(cv, bg=_c("BG"))
    gwin    = cv.create_window((0,0), window=gallery, anchor="nw")

    def upd_sb(*_):
        try:
            t,b=cv.yview(); h=max(sb.winfo_height(),1)
            sb.coords(tr,0,t*h,SB_W,max(b*h,t*h+14))
        except: pass

    def scroll(d): cv.yview_scroll(d,"units"); upd_sb()
    for w in (cv,gallery,sb):
        w.bind("<MouseWheel>", lambda e: scroll(-1*(e.delta//120)))
    def sb_drag(e):
        t,b=cv.yview(); h=max(sb.winfo_height(),1)
        cv.yview_moveto(max(0,e.y/h-(b-t)/2)); upd_sb()
    sb.bind("<ButtonPress-1>",sb_drag); sb.bind("<B1-Motion>",sb_drag)
    gallery.bind("<Configure>", lambda e: [
        cv.configure(scrollregion=cv.bbox("all")), upd_sb()])
    cv.bind("<Configure>", lambda e: [
        cv.itemconfig(gwin,width=e.width), upd_sb()])

    _photos = []

    def render_gallery():
        for w in gallery.winfo_children(): w.destroy()
        _photos.clear()
        chars = list_characters()
        COLS  = 2

        for i, ch in enumerate(chars):
            is_active  = ch["is_active"]
            border_col = _c("ACCENT") if is_active else _c("BORDER")

            cell_b = tk.Frame(gallery, bg=border_col)
            cell_b.grid(row=i//COLS, column=i%COLS,
                        padx=6, pady=6, sticky="nsew")
            cell = tk.Frame(cell_b, bg=_c("CARD"), cursor="hand2")
            cell.pack(fill="both", padx=2, pady=2)

            # Thumbnail from stopped frame
            photo = load_thumbnail(ch["path"], 100)
            if photo:
                _photos.append(photo)
                tk.Label(cell, image=photo, bg=_c("CARD")).pack(pady=(10,4))
            else:
                tk.Label(cell, text="🐱", font=tkfont.Font(size=32),
                         bg=_c("CARD")).pack(pady=(10,4))

            tk.Label(cell, text=ch["name"],
                     font=fbn if is_active else fb,
                     bg=_c("CARD"),
                     fg=_c("ACCENT") if is_active else _c("TEXT")).pack()

            # Theme preview badge — for custom chars show active theme, else mapped theme
            t_key = _theme.theme_for_character(ch["path"])
            if t_key == "default" and not (ch["path"] == config.SPRITESHEET):
                # Custom character — show the currently active theme
                t_key = _theme.name()
                # t_key is now a name string, look up by name
                t_key = next((k for k,v in _theme.THEMES.items()
                              if v["name"] == t_key), "default")
            palette = _theme.THEMES.get(t_key, _theme.THEMES["default"])
            badge_row = tk.Frame(cell, bg=_c("CARD"))
            badge_row.pack(pady=(2,0))
            # Filled colour dot using canvas rectangle
            dot_cv = tk.Canvas(badge_row, width=14, height=14,
                               bg=_c("CARD"), bd=0, highlightthickness=0)
           # dot_cv.pack(side="left", padx=(0,4))
            dot_cv.create_oval(2, 2, 12, 12, fill=palette["ACCENT"], outline="")
            tk.Label(badge_row, text=palette["name"],
                     font=fs, bg=_c("CARD"), fg=_c("SUB")).pack(side="left")

            # Active / Select button
            tk.Button(cell,
                      text="✓ Active" if is_active else "Select",
                      font=fs,
                      bg=_c("ACCENT") if is_active else _c("BORDER"),
                      fg=_c("BG")     if is_active else _c("SUB"),
                      activebackground=_c("ACCENT"), activeforeground=_c("BG"),
                      relief="flat", cursor="hand2", bd=0,
                      command=lambda p=ch["path"],n=ch["name"]: select(p,n)
                      ).pack(pady=(4,10), ipadx=12, ipady=4)

            for w in (cell,):
                w.bind("<Button-1>",
                       lambda e,p=ch["path"],n=ch["name"]: select(p,n))

        for c in range(COLS):
            gallery.columnconfigure(c, weight=1)
        cv.update_idletasks()
        cv.configure(scrollregion=cv.bbox("all"))
        upd_sb()

    def select(path, name):
        set_active_character(path)
        if on_character_change:
            on_character_change(path)
        show_msg(f"✓ {name} is now your companion!")
        render_gallery()

    # ── Bottom bar ────────────────────────────────────────────────────────────
    bot = tk.Frame(parent, bg=_c("BG"))
    bot.pack(fill="x", pady=(8,0))
    tk.Button(bot, text="📁  Open characters folder",
              font=fs, bg=_c("BORDER"), fg=_c("ACCENT"),
              activebackground=_c("ACCENT"), activeforeground=_c("BG"),
              relief="flat", cursor="hand2", bd=0,
              command=lambda: os.startfile(CHARACTERS_DIR)
              ).pack(side="left", ipady=5, ipadx=12)
    tk.Button(bot, text="↺  Refresh",
              font=fs, bg=_c("BORDER"), fg=_c("SUB"),
              activebackground=_c("BORDER"), activeforeground=_c("TEXT"),
              relief="flat", cursor="hand2", bd=0,
              command=render_gallery
              ).pack(side="left", padx=8, ipady=5, ipadx=12)

    render_gallery()

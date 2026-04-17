# ── features/ai_chat.py ───────────────────────────────────────────────────────
# Phase 8: AI Chat — chat with your cat companion powered by Claude API
# - Reads GEMINI_API_KEY from .env
# - Context-aware: knows user name, today's todos, health counts
# - Cat personality — friendly, warm, uses cat puns occasionally
# - Conversation history saved to data/chat_history.xlsx
# - Streams response word by word for a live typing effect

import tkinter as tk
from tkinter import font as tkfont
from datetime import datetime, date
import threading, time, os, json
import urllib.request, urllib.error
import config

import theme as _theme_mod
def _c(k): return _theme_mod.get(k)

_chat_placeholder_shown = [False]   # first-time-only placeholder

CHAT_DATA = os.path.join(config.DATA_DIR, "chat_history.xlsx")

CAT_SYSTEM_PROMPT = """You are a friendly desktop cat companion assistant named Whiskers. \
You are helpful, warm, and occasionally use light cat-themed humour (a paw pun here and there \
is fine, but don't overdo it). You are concise — keep replies under 120 words unless the user \
explicitly asks for more detail. You know the user's context (provided below) and can reference \
it naturally when relevant. Never be condescending. Always be encouraging."""

# ── .env helper ───────────────────────────────────────────────────────────────
def _load_env() -> dict:
    env = {}
    env_path = os.path.join(config.BASE_DIR, ".env")
    if not os.path.exists(env_path):
        return env
    try:
        with open(env_path, "r") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, _, v = line.partition("=")
                env[k.strip()] = v.strip().strip('"').strip("'")
    except Exception as e:
        print(f"[chat] .env error: {e}")
    return env

def _get_api_key() -> str:
    return _load_env().get("GEMINI_API_KEY", "")

# ── Context gathering ─────────────────────────────────────────────────────────
def _build_context(user: dict) -> str:
    """Build a context string from user data, todos, health."""
    lines = [f"User name: {user.get('name','Friend')}",
             f"Date: {datetime.now().strftime('%A, %d %B %Y %H:%M')}"]

    # Pending todos
    try:
        import openpyxl
        if os.path.exists(config.TODO_DATA):
            wb = openpyxl.load_workbook(config.TODO_DATA)
            ws = wb.active
            pending = []
            for i, r in enumerate(ws.iter_rows(values_only=True)):
                if i == 0 or not r[0]: continue
                if not r[4]:  # not done
                    pending.append(str(r[1]))
            if pending:
                lines.append(f"Pending tasks: {', '.join(pending[:5])}")
    except Exception:
        pass

    # Today's health
    try:
        import openpyxl
        from datetime import date as date_
        today = str(date_.today())
        if os.path.exists(config.HEALTH_DATA):
            wb = openpyxl.load_workbook(config.HEALTH_DATA)
            ws = wb.active
            counts = {"water": 0, "break": 0, "stretch": 0}
            for i, r in enumerate(ws.iter_rows(values_only=True)):
                if i == 0 or not r[0]: continue
                if str(r[0]) == today and r[1] in counts:
                    counts[str(r[1])] += 1
            lines.append(
                f"Health today: water {counts['water']}x, "
                f"break {counts['break']}x, stretch {counts['stretch']}x")
    except Exception:
        pass

    return "\n".join(lines)

# ── Chat history Excel ─────────────────────────────────────────────────────────
def _init_chat_wb():
    os.makedirs(config.DATA_DIR, exist_ok=True)
    if not os.path.exists(CHAT_DATA):
        try:
            import openpyxl
            wb = openpyxl.Workbook(); ws = wb.active
            ws.title = "History"
            ws.append(["date", "time", "role", "message"])
            wb.save(CHAT_DATA)
        except Exception as e:
            print(f"[chat] init wb: {e}")

def _save_message(role: str, message: str):
    _init_chat_wb()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(CHAT_DATA)
        ws = wb.active
        ws.append([str(date.today()),
                   datetime.now().strftime("%H:%M"),
                   role, message])
        wb.save(CHAT_DATA)
    except Exception as e:
        print(f"[chat] save: {e}")

def _load_recent_history(n: int = 20) -> list[dict]:
    """Load last n messages for context."""
    _init_chat_wb()
    rows = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(CHAT_DATA)
        ws = wb.active
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not r[2]: continue
            rows.append({"role": r[2], "content": r[3] or ""})
    except Exception as e:
        print(f"[chat] load history: {e}")
    return rows[-n:] if len(rows) > n else rows

# ── Claude API call ────────────────────────────────────────────────────────────
# Models to try in order — all free tier, no billing needed
# Models tried in order — all confirmed free on v1beta (no billing needed)
# gemini-2.0-flash: 15 RPM, 1500 RPD free (confirmed March 2026)
# gemini-2.5-flash: free tier, generous limits
# gemini-1.5-flash-latest: stable alias, always resolves to working version
_GEMINI_MODELS = [
    "gemini-2.0-flash",
    "gemini-2.5-flash",
    "gemini-1.5-flash-latest",
]

def _gemini_url(model: str, api_key: str) -> str:
    env = _load_env()
    base = env.get(
        "GEMINI_API_URL",
        "https://generativelanguage.googleapis.com/v1beta/models"
    ).rstrip("/")
    return f"{base}/{model}:generateContent?key={api_key}"

def _sanitise_history(messages: list[dict]) -> list[dict]:
    """
    Gemini requires strict user/model alternation.
    Fix by merging consecutive same-role messages and ensuring
    the conversation always starts with a user message.
    """
    if not messages:
        return []

    # Map roles
    mapped = []
    for m in messages:
        role = "user" if m["role"] == "user" else "model"
        mapped.append({"role": role, "content": m["content"]})

    # Merge consecutive same-role messages
    merged = [mapped[0]]
    for m in mapped[1:]:
        if m["role"] == merged[-1]["role"]:
            merged[-1]["content"] += "\n" + m["content"]
        else:
            merged.append(m)

    # Must start with user
    if merged[0]["role"] != "user":
        merged = merged[1:]

    # Must end with user (the latest message)
    # Already guaranteed since we always append user then model
    return merged

def _call_gemini(messages: list[dict], system: str, api_key: str) -> str:
    """Call Gemini API with automatic model fallback. Returns reply text."""
    clean = _sanitise_history(messages)
    if not clean:
        return "⚠ No message to send."

    gemini_msgs = [
        {"role": m["role"], "parts": [{"text": m["content"]}]}
        for m in clean
    ]

    payload = json.dumps({
        "system_instruction": {"parts": [{"text": system}]},
        "contents": gemini_msgs,
        "generationConfig": {
            "maxOutputTokens": 512,
            "temperature": 0.7,
        }
    }).encode("utf-8")

    last_error = ""
    for model in _GEMINI_MODELS:
        url = _gemini_url(model, api_key)
        req = urllib.request.Request(
            url, data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        try:
            with urllib.request.urlopen(req, timeout=30) as r:
                data = json.loads(r.read())
                text = data["candidates"][0]["content"]["parts"][0]["text"]
                print(f"[chat] success with model: {model}")
                return text
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8", errors="replace")
            print(f"[chat] {model} HTTP {e.code}: {body}")
            last_error = body
            if e.code == 400:
                return "⚠ Invalid request. Please try again."
            if e.code == 403:
                return "⚠ Invalid API key. Check GEMINI_API_KEY in your .env file."
            if e.code == 429:
                print(f"[chat] rate limited on {model}, trying next…")
                time.sleep(2)
                continue  # try next model silently
            if e.code == 404:
                # Model not available in this region/account — try next
                print(f"[chat] {model} not found, trying next…")
                continue
            last_error = f"HTTP {e.code}"
            continue
        except Exception as e:
            print(f"[chat] {model} error: {e}")
            last_error = str(e)
            continue

    # All models failed
    if "429" in last_error or "RESOURCE_EXHAUSTED" in last_error:
        return ("⚠ Rate limit reached — Gemini free tier is busy.\n"
                "Please wait 60 seconds and try again.")
    if "API_KEY" in last_error.upper() or "403" in last_error:
        return "⚠ Invalid API key. Check GEMINI_API_KEY in your .env file."
    if "404" in last_error:
        return "⚠ Gemini model not available. Check your API key tier."
    return "⚠ Could not reach Gemini. Check your internet connection."

# ── Main view ─────────────────────────────────────────────────────────────────
def build(parent: tk.Frame, root: tk.Tk, user: dict = None):
    if user is None:
        user = {}

    api_key = _get_api_key()
    history = _load_recent_history(20)   # load from Excel on open

    fh   = tkfont.Font(family="Segoe UI", size=11, weight="bold")
    #fb   = tkfont.Font(family="Segoe UI", size=10)
    fs   = tkfont.Font(family="Segoe UI", size=9)
    fbn  = tkfont.Font(family="Segoe UI", size=10, weight="bold")
    fe   = tkfont.Font(family="Segoe UI", size=11)
    fmsg = tkfont.Font(family="Segoe UI", size=10)

    # ── No API key screen ─────────────────────────────────────────────────────
    if not api_key:
        tk.Label(parent, text="🐱  AI Chat Setup",
                 font=fh, bg=_c("BG"), fg=_c("ACCENT")).pack(pady=(30,8))
        tk.Label(parent,
                 text="Add your Anthropic API key to the .env file\n"
                      "in your project folder to enable AI chat.\n\n"
                      "GEMINI_API_KEY=your_key_here\n\n"
                      "Get a FREE key at: aistudio.google.com",
                 font=fs, bg=_c("BG"), fg=_c("SUB"), justify="center").pack()
        tk.Button(parent, text="↺  I've added my key — reload",
                  font=fbn, bg=_c("ACCENT"), fg=_c("BG"),
                  activebackground=_c("TEXT"), activeforeground=_c("BG"),
                  relief="flat", cursor="hand2", bd=0,
                  command=lambda: [w.destroy() for w in parent.winfo_children()] or build(parent, root, user)
                  ).pack(pady=20, ipady=8, ipadx=16)
        return

    # ── Chat layout ───────────────────────────────────────────────────────────
    # Top bar: model info + clear button
    top = tk.Frame(parent, bg=_c("BG"))
    top.pack(fill="x", pady=(0,6))
    tk.Label(top, text="🐱  Whiskers — powered by Gemini 2.0 (free)",
             font=fs, bg=_c("BG"), fg=_c("SUB")).pack(side="left")
    tk.Button(top, text="Clear chat",
              font=fs, bg=_c("BG"), fg=_c("SUB"),
              activebackground=_c("CARD"), activeforeground=_c("ACCENT"),
              relief="flat", cursor="hand2", bd=0,
              command=lambda: clear_chat()
              ).pack(side="right")

    # Message list — scrollable canvas
    msg_outer = tk.Frame(parent, bg=_c("BG"))
    msg_outer.pack(fill="both", expand=True, pady=(0,6))

    cv = tk.Canvas(msg_outer, bg=_c("BG"), bd=0, highlightthickness=0)
    cv.pack(side="left", fill="both", expand=True)

    SB_W = 4
    sb = tk.Canvas(msg_outer, width=SB_W, bg=_c("BG"), bd=0, highlightthickness=0)
    sb.pack(side="right", fill="y", padx=(2,0))
    thumb = sb.create_rectangle(0,0,SB_W,30, fill=_c("ACCENT"), outline="")

    msg_frame = tk.Frame(cv, bg=_c("BG"))
    mwin = cv.create_window((0,0), window=msg_frame, anchor="nw")

    def update_thumb(*_):
        try:
            top_f, bot_f = cv.yview()
            h = max(sb.winfo_height(), 1)
            sb.coords(thumb, 0, top_f*h, SB_W, max(bot_f*h, top_f*h+14))
        except: pass

    def scroll(delta):
        top, bot = cv.yview()
        if top == 0.0 and bot == 1.0: return
        cv.yview_scroll(delta, "units")
        update_thumb()

    for w in (cv, msg_frame, sb):
        w.bind("<MouseWheel>", lambda e: scroll(-1*(e.delta//120)))

    def sb_drag(e):
        top_f, bot_f = cv.yview()
        h = max(sb.winfo_height(), 1)
        cv.yview_moveto(max(0, e.y/h-(bot_f-top_f)/2))
        update_thumb()

    sb.bind("<ButtonPress-1>", sb_drag)
    sb.bind("<B1-Motion>",     sb_drag)

    def _on_msg_configure(e):
        bb = cv.bbox("all")
        if bb: cv.configure(scrollregion=(0, 0, bb[2], bb[3]))
        update_thumb()
    msg_frame.bind("<Configure>", _on_msg_configure)
    cv.yview_moveto(0)   # start at top
    cv.bind("<Configure>", lambda e: [
        cv.itemconfig(mwin, width=e.width), update_thumb()])

    # ── Message bubble builder ────────────────────────────────────────────────
    def add_bubble(role: str, text: str) -> tk.Label:
        """Add a message bubble and return the label (so we can update it for streaming)."""
        is_user = (role == "user")

        row = tk.Frame(msg_frame, bg=_c("BG"))
        row.pack(fill="x", pady=3, padx=6)

        bubble_bg  = _c("ACCENT") if is_user else _c("CARD")
        bubble_fg  = _c("BG") if is_user else _c("TEXT")
        align      = "e" if is_user else "w"
        padx_side  = (60, 0) if is_user else (0, 60)

        bubble = tk.Frame(row, bg=bubble_bg)
        bubble.pack(anchor=align, padx=padx_side)

        lbl = tk.Label(bubble, text=text,
                       font=fmsg, bg=bubble_bg, fg=bubble_fg,
                       wraplength=300, justify="left" if not is_user else "right",
                       padx=12, pady=8)
        lbl.pack()

        # Bind mouse wheel on every bubble so scroll works anywhere in chat
        for w in (row, bubble, lbl):
            w.bind("<MouseWheel>", lambda e: scroll(-1*(e.delta//120)))

        # Scroll to bottom — let scrollregion update before yview
        def _scroll_bottom():
            bb = cv.bbox("all")
            if bb: cv.configure(scrollregion=(0, 0, bb[2], bb[3]))
            cv.yview_moveto(1.0)
            update_thumb()
        root.after_idle(lambda: root.after(80, _scroll_bottom))
        return lbl

    def add_typing_indicator() -> tk.Label:
        """Shows '...' while waiting for response."""
        return add_bubble("assistant", "●  ●  ●")

    # ── Render history on load ────────────────────────────────────────────────
    if history:
        for msg in history:
            add_bubble(msg["role"], msg["content"])
    else:
        # Welcome message
        name = user.get("name", "there")
        add_bubble("assistant",
                   f"Hey {name}! 🐱 I'm Whiskers, your desktop companion. "
                   f"Ask me anything — I know your tasks and health status too!")

    # ── Input bar ─────────────────────────────────────────────────────────────
    input_row = tk.Frame(parent, bg=_c("BG"))
    input_row.pack(fill="x")

    entry_wrap = tk.Frame(input_row, bg=_c("BORDER"))
    entry_wrap.pack(side="left", fill="x", expand=True, padx=(0,6))

    entry = tk.Entry(entry_wrap,
                     font=fe, bg=_c("CARD"), fg=_c("SUB"),
                     insertbackground=_c("ACCENT"),
                     relief="flat", bd=0,
                     highlightthickness=0)
    entry.pack(fill="x", padx=1, pady=1, ipady=9)

    if not _chat_placeholder_shown[0]:
        entry.insert(0, "Ask Whiskers anything…")

    def _clear_placeholder(e=None):
        if not _chat_placeholder_shown[0]:
            entry.delete(0, "end")
            entry.config(fg=_c("TEXT"))
            _chat_placeholder_shown[0] = True

    entry.bind("<FocusIn>",  _clear_placeholder)
    entry.bind("<Button-1>", _clear_placeholder)
    entry.focus()

    send_btn = tk.Button(input_row, text="Send  ▶",
                         font=fbn, bg=_c("ACCENT"), fg=_c("BG"),
                         activebackground=_c("TEXT"), activeforeground=_c("BG"),
                         relief="flat", cursor="hand2", bd=0,
                         command=lambda: send_message())
    send_btn.pack(side="left", ipady=9, ipadx=14)
    entry.bind("<Return>", lambda e: send_message())

    # ── Send logic ────────────────────────────────────────────────────────────
    def send_message():
        text = entry.get().strip()
        if not text: return

        entry.delete(0, "end")

        # Show user bubble
        add_bubble("user", text)
        history.append({"role":"user","content":text})
        _save_message("user", text)

        # Disable send while waiting
        send_btn.config(state="disabled", bg=_c("BORDER"), fg=_c("SUB"))

        # Show typing indicator
        typing_lbl = add_typing_indicator()

        def fetch():
            # Build system prompt with live context
            context = _build_context(user)
            system  = f"{CAT_SYSTEM_PROMPT}\n\nUser context:\n{context}"

            reply = _call_gemini(history, system, api_key)

            def on_reply():
                # Guard: if user navigated away, all widgets are destroyed — bail out
                try:
                    if not msg_frame.winfo_exists():
                        return
                except Exception:
                    return

                # Remove typing indicator row
                try:
                    typing_lbl.master.master.destroy()
                except Exception:
                    pass

                # Add real reply bubble
                add_bubble("assistant", reply)
                # Only add to history if it's a real reply, not an error
                if not reply.startswith("⚠"):
                    history.append({"role":"assistant","content":reply})
                    _save_message("assistant", reply)

                # Re-enable send (no placeholder — first-time-only behaviour)
                try:
                    send_btn.config(state="normal", bg=_c("ACCENT"), fg=_c("BG"))
                    entry.focus()
                except Exception:
                    pass

            try:
                root.after(0, on_reply)
            except Exception:
                pass

        threading.Thread(target=fetch, daemon=True).start()

    def clear_chat():
        # Clear in-memory history
        history.clear()
        # Clear saved Excel history so reopen shows clean slate
        try:
            import openpyxl
            os.makedirs(config.DATA_DIR, exist_ok=True)
            wb = openpyxl.Workbook(); ws = wb.active
            ws.title = "History"
            ws.append(["date","time","role","message"])
            wb.save(CHAT_DATA)
        except Exception as e:
            print(f"[chat] clear history: {e}")
        for w in msg_frame.winfo_children(): w.destroy()
        name = user.get("name","there")
        add_bubble("assistant",
                   f"Chat cleared! 🐱 What\'s on your mind, {name}?")

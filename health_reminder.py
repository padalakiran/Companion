# ── health_reminder.py ────────────────────────────────────────────────────────
# Fixes:
#  1. threading.Lock on Excel writes — no more corruption
#  2. Popup queue — multiple simultaneous reminders show one after another
#  3. Per-reminder enable/disable (saved to user.xlsx)

import tkinter as tk
from tkinter import font as tkfont
import threading, time, os, queue
from datetime import datetime, date
import config

# Colours read live inside _show() via theme.get() — no frozen vars needed

# ── Write lock shared across all threads ─────────────────────────────────────
_excel_lock = threading.Lock()

# ── Cat speech-bubble popup ───────────────────────────────────────────────────
class _CatPopup:
    """
    Shows one popup at a time from a queue.
    When the current popup dismisses, the next queued one appears.
    """
    W = 320; H = 160

    def __init__(self, root: tk.Tk):
        self._root     = root
        self._win      = None
        self._after_id = None
        self._photo    = None
        self._queue: queue.Queue = queue.Queue()
        self._showing  = False
        self._load_image()

    def _load_image(self):
        """Load character image — reloads if character OR theme changes."""
        try:
            from PIL import Image, ImageTk
            from features.character import get_active_character
            active_path = get_active_character()
            sheet = Image.open(active_path).convert("RGBA")
            w, h  = sheet.size
            S     = config.SPRITE_SIZE
            # Check if it's a proper spritesheet (5 rows × 4 cols)
            if h >= S * 5 and w >= S * 4:
                # Use stopped frame (row 4, col 0) — character facing forward
                frame = sheet.crop((0, 4*S, S, 5*S))
            else:
                # Static image — use whole image
                sq = min(w, h)
                frame = sheet.crop(((w-sq)//2, (h-sq)//2, (w+sq)//2, (h+sq)//2))
            frame = frame.resize((72, 72), Image.LANCZOS)
            # Place on theme-coloured background
            import theme as _t
            # Parse bg colour from hex
            bg_hex = _t.get("CARD").lstrip("#")
            bg_r, bg_g, bg_b = int(bg_hex[0:2],16), int(bg_hex[2:4],16), int(bg_hex[4:6],16)
            bg    = Image.new("RGBA", (72, 72), (bg_r, bg_g, bg_b, 255))
            bg.paste(frame, (0, 0), frame)
            self._photo = ImageTk.PhotoImage(bg)
            self._photo_path = active_path  # track which character is loaded
        except Exception as e:
            print(f"[popup] image: {e}")
            self._photo = None

    def enqueue(self, title: str, message: str, color: str):
        """Thread-safe: add a popup to the queue."""
        self._queue.put((title, message, color))
        # Schedule showing on main thread
        try:
            self._root.after(0, self._try_show_next)
        except Exception:
            pass

    def _try_show_next(self):
        """Show next queued popup if none is currently visible."""
        if self._showing:
            return
        try:
            title, message, color = self._queue.get_nowait()
        except queue.Empty:
            return
        self._show(title, message, color)

    def _show(self, title: str, message: str, color: str):
        self._showing = True
        if self._after_id:
            try: self._root.after_cancel(self._after_id)
            except: pass

        try:
            if self._win and self._win.winfo_exists():
                self._win.destroy()
        except: pass

        # Read live theme colours every time popup is shown
        import theme as _t
        _BG = _t.get('BG'); _CARD = _t.get('CARD'); _BORDER = _t.get('BORDER')
        _TEXT = _t.get('TEXT'); _ACCENT = _t.get('ACCENT'); _SUB = _t.get('SUB')

        # Always reload image — ensures theme bg is always current
        self._load_image()

        try:
            sw = self._root.winfo_screenwidth()
            sh = self._root.winfo_screenheight()

            win = tk.Toplevel(self._root)
            win.overrideredirect(True)
            win.attributes("-topmost", True)
            win.geometry(f"{self.W}x{self.H}+{sw-self.W-20}+{sh-self.H-80}")
            win.config(highlightthickness=0, bd=0)
            win.configure(bg=_CARD)
            # Suppress Windows DWM system border/shadow entirely
            try:
                import ctypes
                ctypes.windll.dwmapi.DwmSetWindowAttribute(
                    ctypes.windll.user32.GetParent(win.winfo_id()),
                    2, ctypes.byref(ctypes.c_int(1)), ctypes.sizeof(ctypes.c_int))
            except Exception: pass
            self._win = win

            inner = tk.Frame(win, bg=_CARD, bd=0, highlightthickness=0)
            inner.place(x=0, y=0, width=self.W, height=self.H)

            if self._photo:
                tk.Label(inner, image=self._photo, bg=_CARD, cursor="hand2").place(
                    x=10, y=(self.H-72)//2)

            bx = 92; bw = self.W-bx-10
            bubble = tk.Frame(inner, bg=_BORDER, bd=0)
            bubble.place(x=bx, y=10, width=bw, height=self.H-20)
            bi = tk.Frame(bubble, bg=_BG, bd=0)
            bi.place(x=1, y=1, width=bw-2, height=self.H-22)

            t1 = tk.Canvas(inner, width=10, height=14, bg=_CARD,
                           bd=0, highlightthickness=0)
            t1.place(x=bx-10, y=32)
            t1.create_polygon(10,0, 10,14, 0,7, fill=_BORDER, outline="")
            t2 = tk.Canvas(inner, width=8, height=12, bg=_CARD,
                           bd=0, highlightthickness=0)
            t2.place(x=bx-8, y=33)
            t2.create_polygon(8,0, 8,12, 0,6, fill=_BG, outline="")

            ft = tkfont.Font(family="Segoe UI", size=10, weight="bold")
            fm = tkfont.Font(family="Segoe UI", size=9)
            fb = tkfont.Font(family="Segoe UI", size=8)

            tk.Label(bi, text=title,   font=ft, bg=_BG, fg=_ACCENT,
                     anchor="w").place(x=10, y=10, width=bw-22)
            tk.Label(bi, text=message, font=fm, bg=_BG, fg=_TEXT,
                     anchor="w", wraplength=bw-22,
                     justify="left").place(x=10, y=32, width=bw-22)

            def dismiss():
                self._showing = False
                try:
                    if self._after_id:
                        self._root.after_cancel(self._after_id)
                except: pass
                try:
                    if win.winfo_exists(): win.destroy()
                except: pass
                # Show next queued popup after short delay
                self._root.after(400, self._try_show_next)

            tk.Button(bi, text="OK  ✓", font=fb, bg=_BORDER, fg=_TEXT,
                      activebackground=color, activeforeground=_BG,
                      relief="flat", cursor="hand2", bd=0,
                      command=dismiss).place(x=bw-60,
                                             y=self.H-4-30-22,
                                             width=50, height=20)
            for w in (win, inner, bi):
                w.bind("<Button-1>", lambda e: dismiss())

            self._after_id = self._root.after(7000, dismiss)

        except Exception as e:
            print(f"[popup] show error: {e}")
            self._showing = False
            self._root.after(500, self._try_show_next)


# ── Health Reminder engine ────────────────────────────────────────────────────
class HealthReminder:
    @property
    def REMINDERS(self):
        icon = __import__('theme').icon()
        return {
            "water":   (f"{icon} Time to drink water!",  "Stay hydrated — grab a glass of water.", "#4FC3F7"),
            "break":   (f"{icon} Short break time!",     "Rest your eyes for a few minutes.",      "#FFB74D"),
            "stretch": (f"{icon} Time to stretch!",      "Stand up and stretch for 2 minutes.",    "#81C784"),
        }

    def __init__(self, root: tk.Tk):
        self._root    = root
        self._popup   = _CatPopup(root)
        self._running = False
        self._thread  = None
        self._ui_cb   = None

        # Countdown in seconds
        self._countdown = {
            "water":   self._load_interval("water_interval",   config.WATER_INTERVAL_MIN)   * 60,
            "break":   self._load_interval("break_interval",   config.BREAK_INTERVAL_MIN)   * 60,
            "stretch": self._load_interval("stretch_interval", config.STRETCH_INTERVAL_MIN) * 60,
        }
        self._interval = dict(self._countdown)

        # Per-reminder enabled/disabled
        self._enabled = {
            "water":   self._load_enabled("water_enabled",   True),
            "break":   self._load_enabled("break_enabled",   True),
            "stretch": self._load_enabled("stretch_enabled", True),
        }

    # ── Public ────────────────────────────────────────────────────────────────

    def start(self):
        if self._running and self._thread and self._thread.is_alive():
            return
        self._running = True
        self._thread  = threading.Thread(target=self._loop, daemon=True)
        self._thread.start()
        print(f"[health] started — "
              f"water={self._interval['water']//60}m "
              f"break={self._interval['break']//60}m "
              f"stretch={self._interval['stretch']//60}m")

    def stop(self):
        self._running = False

    def set_interval(self, key: str, minutes: int):
        secs = minutes * 60
        self._countdown[key] = secs
        self._interval[key]  = secs
        attr = {"water":"WATER_INTERVAL_MIN",
                "break":"BREAK_INTERVAL_MIN",
                "stretch":"STRETCH_INTERVAL_MIN"}.get(key)
        if attr:
            setattr(config, attr, minutes)

    def set_enabled(self, key: str, enabled: bool):
        """Enable or disable a specific reminder."""
        self._enabled[key] = enabled
        self._save_user(f"{key}_enabled", "1" if enabled else "0")
        print(f"[health] {key} → {'enabled' if enabled else 'disabled'}")

    def is_enabled(self, key: str) -> bool:
        return self._enabled.get(key, True)

    def register_ui(self, fn):
        self._ui_cb = fn

    def unregister_ui(self):
        self._ui_cb = None

    # ── Loop ──────────────────────────────────────────────────────────────────

    def _loop(self):
        while self._running:
            try:
                time.sleep(1)
                if not self._running: break

                for key in ("water", "break", "stretch"):
                    try:
                        self._countdown[key] -= 1
                        if self._countdown[key] <= 0:
                            self._countdown[key] = self._interval[key]
                            if self._enabled.get(key, True):
                                threading.Thread(
                                    target=self._fire,
                                    args=(key,),
                                    daemon=True
                                ).start()
                    except Exception as e:
                        print(f"[health] key={key}: {e}")
                        self._countdown[key] = self._interval.get(key, 3600)

            except Exception as e:
                print(f"[health] loop error: {e}")
                time.sleep(1)

        print("[health] engine stopped")

    def _fire(self, key: str):
        title, message, color = self.REMINDERS[key]
        print(f"[health] 🔔 {key} at {datetime.now().strftime('%H:%M:%S')}")

        # Log with lock — prevents concurrent write corruption
        try:
            self._log(key)
        except Exception as e:
            print(f"[health] log error (non-fatal): {e}")

        # Enqueue popup (queue handles serialisation)
        try:
            self._root.after(0, lambda t=title, m=message, c=color:
                             self._popup.enqueue(t, m, c))
        except Exception as e:
            print(f"[health] popup error: {e}")

        # Update dashboard counts if open
        if self._ui_cb:
            try:
                self._root.after(0, lambda k=key: self._ui_cb(k))
            except Exception:
                pass

    def _log(self, key: str):
        """Write to Excel with a global lock to prevent concurrent corruption."""
        import openpyxl
        path = config.HEALTH_DATA
        os.makedirs(config.DATA_DIR, exist_ok=True)
        with _excel_lock:
            if os.path.exists(path):
                wb = openpyxl.load_workbook(path)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Health Log"
                ws.append(["date", "type", "time"])
            ws.append([str(date.today()), key,
                       datetime.now().strftime("%H:%M")])
            wb.save(path)

    # ── Persistence helpers ───────────────────────────────────────────────────

    def _load_interval(self, key: str, default: int) -> int:
        try:
            import openpyxl
            if os.path.exists(config.USER_DATA):
                wb = openpyxl.load_workbook(config.USER_DATA)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    if row[0] == key:
                        return int(row[1])
        except Exception:
            pass
        return default

    def _load_enabled(self, key: str, default: bool) -> bool:
        try:
            import openpyxl
            if os.path.exists(config.USER_DATA):
                wb = openpyxl.load_workbook(config.USER_DATA)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    if row[0] == key:
                        return str(row[1]) != "0"
        except Exception:
            pass
        return default

    def _save_user(self, key: str, value: str):
        import openpyxl
        os.makedirs(config.DATA_DIR, exist_ok=True)
        with _excel_lock:
            try:
                if os.path.exists(config.USER_DATA):
                    wb = openpyxl.load_workbook(config.USER_DATA)
                    ws = wb.active
                    for row in ws.iter_rows():
                        if row[0].value == key:
                            row[1].value = value
                            wb.save(config.USER_DATA)
                            return
                    ws.append([key, value])
                    wb.save(config.USER_DATA)
                else:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "User"
                    ws.append(["key", "value"])
                    ws.append([key, value])
                    wb.save(config.USER_DATA)
            except Exception as e:
                print(f"[health] save_user error: {e}")

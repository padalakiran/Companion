# ── user_data.py ──────────────────────────────────────────────────────────────
# Handles first-launch name prompt and persisting user settings to Excel.
# FIX: name dialog uses tk.Toplevel (not a separate tk.Tk) so the main loop
#      stays intact. Data dir is created before saving to avoid silent failures.

import os
import tkinter as tk
from tkinter import font as tkfont
import openpyxl
import config


def get_or_create_user(root: tk.Tk) -> dict:
    """
    Returns user dict with at least {"name": str}.
    First launch  → show name dialog (Toplevel on root) → save → return.
    Later launches → load from Excel → return immediately.
    """
    os.makedirs(config.DATA_DIR, exist_ok=True)   # ensure data/ exists

    if os.path.exists(config.USER_DATA):
        data = _load_user()
        if data.get("name"):          # valid saved name
            return data

    # First launch (or corrupted file) → ask for name
    name = _prompt_name(root)
    user = {"name": name}
    _save_user(user)
    return user


# ── Excel helpers ──────────────────────────────────────────────────────────────

def _load_user() -> dict:
    try:
        wb = openpyxl.load_workbook(config.USER_DATA)
        ws = wb.active
        data = {}
        for row in ws.iter_rows(values_only=True):
            if row[0] and row[1] and row[0] != "key":
                data[str(row[0])] = str(row[1])
        return data
    except Exception as e:
        print(f"[user_data] Load error: {e}")
        return {}


def _save_user(data: dict):
    os.makedirs(config.DATA_DIR, exist_ok=True)
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "User"
        ws.append(["key", "value"])
        for k, v in data.items():
            ws.append([k, v])
        wb.save(config.USER_DATA)
        print(f"[user_data] Saved to {config.USER_DATA}")
    except Exception as e:
        print(f"[user_data] Save error: {e}")


def get_installed_version() -> str:
    """Read app_version from user.xlsx. Returns empty string if not set."""
    try:
        data = _load_user()
        return str(data.get("app_version", "")).strip()
    except Exception:
        return ""


def save_installed_version(version: str):
    """Write app_version into user.xlsx (adds/updates the key)."""
    try:
        data = _load_user()
        data["app_version"] = str(version).strip()
        _save_user(data)
        print(f"[user_data] app_version saved: {version}")
    except Exception as e:
        print(f"[user_data] save_installed_version error: {e}")


# ── Welcome dialog ───────────────────────────────────────────────────────────

def _prompt_name(root: tk.Tk) -> str:
    """
    Custom themed welcome window shown only on first launch.
    Frameless, centred, uses live theme colours.
    Blocks via wait_window() — no separate mainloop.
    """
    import theme as _t

    result = {"name": ""}

    BG     = _t.get("BG")
    CARD   = _t.get("CARD")
    BORDER = _t.get("BORDER")
    ACCENT = _t.get("ACCENT")
    TEXT   = _t.get("TEXT")
    SUB    = _t.get("SUB")

    W, H = 380, 310
    sw   = root.winfo_screenwidth()
    sh   = root.winfo_screenheight()

    dlg = tk.Toplevel(root)
    dlg.overrideredirect(True)
    dlg.attributes("-topmost", True)
    dlg.configure(bg=ACCENT)
    dlg.geometry(f"{W}x{H}+{(sw-W)//2}+{(sh-H)//2}")
    dlg.grab_set()
    dlg.focus_force()

    # Suppress Windows DWM border
    try:
        import ctypes
        ctypes.windll.dwmapi.DwmSetWindowAttribute(
            ctypes.windll.user32.GetParent(dlg.winfo_id()),
            2, ctypes.byref(ctypes.c_int(1)), ctypes.sizeof(ctypes.c_int))
    except Exception: pass

    # Drag support
    _drag = {"x": 0, "y": 0}
    def drag_start(e): _drag["x"] = e.x_root - dlg.winfo_x(); _drag["y"] = e.y_root - dlg.winfo_y()
    def drag_move(e):  dlg.geometry(f"+{e.x_root - _drag['x']}+{e.y_root - _drag['y']}")

    # ── Outer accent border ───────────────────────────────────────────────────
    outer = tk.Frame(dlg, bg=ACCENT, bd=0)
    outer.place(x=0, y=0, width=W, height=H)

    inner = tk.Frame(outer, bg=BG, bd=0)
    inner.place(x=1, y=1, width=W-2, height=H-2)

    # ── Custom titlebar ───────────────────────────────────────────────────────
    titlebar = tk.Frame(inner, bg=CARD, height=32)
    titlebar.pack(fill="x")
    titlebar.pack_propagate(False)

    tk.Label(titlebar, text="  Desktop Companion",
             font=tkfont.Font(family="Segoe UI", size=9),
             bg=CARD, fg=SUB, anchor="w").pack(side="left", fill="y")

    # Minimise button
    def do_minimise():
        dlg.grab_release()   # release grab BEFORE withdraw — otherwise rb can't receive clicks
        dlg.withdraw()
        # Small restore button near taskbar
        rb = tk.Toplevel(root)
        rb.overrideredirect(True)
        rb.attributes("-topmost", True)
        rb.geometry(f"44x28+{sw-50}+{sh-80}")
        rb.configure(bg=CARD)
        def restore():
            rb.destroy()
            dlg.deiconify()
            dlg.focus_force()
            dlg.grab_set()   # re-acquire grab after restore
        tk.Button(rb, text=_t.icon(),
                  font=tkfont.Font(family="Segoe UI Emoji", size=13),
                  bg=CARD, fg=ACCENT, relief="flat", cursor="hand2", bd=0,
                  command=restore).pack(fill="both", expand=True)

    def do_close():
        result["name"] = "Friend"
        dlg.destroy()

    f_ctrl = tkfont.Font(family="Segoe UI", size=10)
    close_btn = tk.Button(titlebar, text="✕", font=f_ctrl,
                          bg=CARD, fg=SUB, activebackground="#F38BA8",
                          activeforeground=BG, relief="flat", cursor="hand2",
                          bd=0, padx=10, command=do_close)
    close_btn.pack(side="right", fill="y")

    min_btn = tk.Button(titlebar, text="─", font=f_ctrl,
                        bg=CARD, fg=SUB, activebackground=BORDER,
                        activeforeground=TEXT, relief="flat", cursor="hand2",
                        bd=0, padx=10, command=do_minimise)
    min_btn.pack(side="right", fill="y")

    # Make titlebar draggable
    for w in (titlebar,):
        w.bind("<ButtonPress-1>",  drag_start)
        w.bind("<B1-Motion>",      drag_move)

    # Separator line under titlebar
    tk.Frame(inner, bg=BORDER, height=1).pack(fill="x")

    # ── Content area ──────────────────────────────────────────────────────────
    content = tk.Frame(inner, bg=BG)
    content.pack(fill="both", expand=True)

    # Icon
    tk.Label(content, text=_t.icon(),
             font=tkfont.Font(family="Segoe UI Emoji", size=38),
             bg=BG, fg=ACCENT).pack(pady=(18, 0))

    # Title
    tk.Label(content, text="Welcome!",
             font=tkfont.Font(family="Segoe UI", size=14, weight="bold"),
             bg=BG, fg=ACCENT).pack(pady=(6, 2))

    tk.Label(content, text="What should your companion call you?",
             font=tkfont.Font(family="Segoe UI", size=9),
             bg=BG, fg=SUB).pack()

    # Entry
    entry_border = tk.Frame(content, bg=BORDER)
    entry_border.pack(padx=36, pady=(12, 0), fill="x")

    entry = tk.Entry(entry_border,
                     font=tkfont.Font(family="Segoe UI", size=12),
                     bg=CARD, fg=TEXT,
                     insertbackground=ACCENT,
                     relief="flat", bd=0,
                     justify="center",
                     highlightthickness=0)
    entry.pack(fill="x", padx=1, pady=1, ipady=8)
    entry.focus_set()

    def on_focus_in(e):  entry_border.config(bg=ACCENT)
    def on_focus_out(e): entry_border.config(bg=BORDER)
    entry.bind("<FocusIn>",  on_focus_in)
    entry.bind("<FocusOut>", on_focus_out)

    # Error label
    err_lbl = tk.Label(content, text="",
                       font=tkfont.Font(family="Segoe UI", size=9),
                       bg=BG, fg="#F38BA8")
    err_lbl.pack(pady=(4, 0))

    # Confirm button
    def confirm(*_):
        n = entry.get().strip()
        if not n:
            err_lbl.config(text="Please enter your name to continue")
            entry_border.config(bg="#F38BA8")
            entry.focus_set()
            return
        result["name"] = n
        dlg.destroy()

    btn = tk.Button(content, text=f"Let's Go  {_t.icon()}",
                    font=tkfont.Font(family="Segoe UI", size=11, weight="bold"),
                    bg=ACCENT, fg=BG,
                    activebackground=TEXT, activeforeground=BG,
                    relief="flat", cursor="hand2", bd=0,
                    command=confirm)
    btn.pack(padx=36, pady=(6, 0), fill="x", ipady=8)
    btn.bind("<Enter>", lambda e: btn.config(bg=TEXT))
    btn.bind("<Leave>", lambda e: btn.config(bg=ACCENT))

    tk.Label(content, text="You can change this anytime in Settings",
             font=tkfont.Font(family="Segoe UI", size=8),
             bg=BG, fg=SUB).pack(pady=(8, 0))

    dlg.bind("<Return>", confirm)

    root.wait_window(dlg)

    return result["name"] if result["name"] else "Friend"

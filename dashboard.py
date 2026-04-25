import tkinter as tk
from tkinter import font as tkfont
from datetime  import datetime, date
#import threading
import time
import os
import openpyxl
import config
import theme
# ── Palette — always read live from theme ─────────────────────────────────────
import theme as _theme_mod

# Module-level colour vars — refreshed by _refresh_colours() before every build
BG=CARD=BORDER=ACCENT=TEXT=SUB=SUCCESS=DANGER=WARN=""

def _refresh_colours():
    global BG,CARD,BORDER,ACCENT,TEXT,SUB,SUCCESS,DANGER,WARN
    BG      = _theme_mod.get("BG")
    CARD    = _theme_mod.get("CARD")
    BORDER  = _theme_mod.get("BORDER")
    ACCENT  = _theme_mod.get("ACCENT")
    TEXT    = _theme_mod.get("TEXT")
    SUB     = _theme_mod.get("SUB")
    SUCCESS = _theme_mod.get("SUCCESS")
    DANGER  = _theme_mod.get("DANGER")
    WARN    = _theme_mod.get("YELLOW")

_refresh_colours()   # initialise on import

# ── Tiles ─────────────────────────────────────────────────────────────────────
TILES = [
    {"key": "health",   "label": "Health",        "sub": "Reminders & log",   "color": "#4FC3A1"},
    {"key": "planner",  "label": "Daily Planner",  "sub": "Today's schedule",  "color": "#4A90D9"},
    {"key": "notes",    "label": "Notes",           "sub": "Quick notes",       "color": "#7B68EE"},
    {"key": "weather",  "label": "Weather",         "sub": "Bengaluru, IN",     "color": "#5BA85B", "sub_key": "weather_city"},
    {"key": "pomodoro", "label": "Pomodoro",        "sub": "Focus timer",       "color": "#E07B54"},
    {"key": "todo",     "label": "To-Do List",      "sub": "Tasks & reminders", "color": "#D4A843"},
    {"key": "ai_chat",  "label": "AI Chat",          "sub": "Chat with Whiskers", "color": "#CBA6F7"},
    {"key": "settings",   "label": "Settings",          "sub": "Preferences",          "color": "#888888"},
    {"key": "character",  "label": "Character",         "sub": "Choose companion",      "color": "#E07B9A"},
]

# ── Icon drawing ──────────────────────────────────────────────────────────────
def draw_icon(canvas, key, cx, cy, size, color):
    import math
    h = size // 2
    if key == "health":
        canvas.create_oval(cx-10, cy-14, cx+10, cy+6,  fill=color, outline="")
        canvas.create_polygon(cx-10,cy-4, cx,cy+14, cx+10,cy-4, fill=color, outline="")
        # Use contrasting color for inner detail so it's visible on hover
        inner_col = "#1A1B2E" if color == "white" else "white"
        canvas.create_oval(cx-5, cy-10, cx+5, cy-2, fill=inner_col, outline="")
    elif key == "planner":
        canvas.create_rectangle(cx-h,cy-h+4,cx+h,cy+h, outline=color,width=2,fill="")
        canvas.create_line(cx-h,cy-h+10,cx+h,cy-h+10, fill=color,width=2)
        canvas.create_line(cx-h+6,cy-h,cx-h+6,cy-h+8, fill=color,width=2)
        canvas.create_line(cx+h-6,cy-h,cx+h-6,cy-h+8, fill=color,width=2)
        for r in range(2):
            for c in range(3):
                canvas.create_oval(cx-10+c*10-2,cy-2+r*9-2,cx-10+c*10+2,cy-2+r*9+2,fill=color,outline="")
    elif key == "notes":
        canvas.create_rectangle(cx-h+2,cy-h,cx+h-2,cy+h, outline=color,width=2,fill="")
        for i in range(3):
            canvas.create_line(cx-h+8,cy-8+i*8,cx+h-8,cy-8+i*8, fill=color,width=2)
    elif key == "weather":
        canvas.create_oval(cx-8,cy-12,cx+8,cy+4,   fill=color,outline="")
        canvas.create_oval(cx-14,cy-6,cx+2,cy+6,   fill=color,outline="")
        canvas.create_oval(cx-4,cy-4,cx+14,cy+8,   fill=color,outline="")
        canvas.create_oval(cx+4,cy-16,cx+12,cy-8,  fill="#F5C842",outline="")
    elif key == "pomodoro":
        canvas.create_oval(cx-h+2,cy-h+6,cx+h-2,cy+h, fill=color,outline=color)
        canvas.create_line(cx,cy-h+6,cx,cy-h, fill=color,width=3)
        canvas.create_line(cx,cy-h,cx+6,cy-h-4, fill="#5BA85B",width=2)
        hand_col = "#1A1B2E" if color == "white" else "white"
        canvas.create_line(cx,cy-2,cx,cy+8,   fill=hand_col,width=2)
        canvas.create_line(cx,cy-2,cx+7,cy+2, fill=hand_col,width=2)
    elif key == "todo":
        canvas.create_rectangle(cx-h,cy-h,cx+h,cy+h, outline=color,width=2,fill="")
        canvas.create_line(cx-8,cy,cx-2,cy+8,cx+10,cy-8, fill=color,width=3,joinstyle="round")
    elif key == "settings":
        canvas.create_oval(cx-8,cy-8,cx+8,cy+8, outline=color,width=2,fill="")
        canvas.create_oval(cx-4,cy-4,cx+4,cy+4, fill=color,outline="")
        for ad in range(0,360,45):
            a=math.radians(ad)
            canvas.create_line(cx+int(8*math.cos(a)),cy+int(8*math.sin(a)),
                               cx+int(13*math.cos(a)),cy+int(13*math.sin(a)),fill=color,width=3)
    elif key == "character":
        # Paw print icon
        canvas.create_oval(cx-h+2,cy-h+2,cx+h-2,cy+h-2, outline=color, width=2, fill="")
        for ax,ay in [(-6,-6),(6,-6),(-8,2),(8,2)]:
            canvas.create_oval(cx+ax-4,cy+ay-4,cx+ax+4,cy+ay+4, fill=color, outline="")
    elif key == "ai_chat":
        # Speech bubble icon
        canvas.create_oval(cx-h,cy-h,cx+h,cy+4, outline=color,width=2,fill="")
        canvas.create_polygon(cx-4,cy+4, cx-10,cy+h, cx+4,cy+4,
                              outline=color,fill=color)
        for dx in [-6,0,6]:
            canvas.create_oval(cx+dx-2,cy-4,cx+dx+2,cy,fill=color,outline="")
    elif key == "water":
        canvas.create_polygon(cx,cy-h, cx+h,cy+4, cx-h,cy+4, fill=color,outline="")
        canvas.create_oval(cx-h,cy+2,cx+h,cy+h, fill=color,outline="")
    elif key == "break_":
        canvas.create_arc(cx-h,cy-h,cx+h,cy+h, start=30,extent=300,
                          outline=color,width=2,style="arc")
        canvas.create_line(cx,cy+h,cx,cy+h+6, fill=color,width=2)
        canvas.create_line(cx-6,cy+h+6,cx+6,cy+h+6, fill=color,width=2)
    elif key == "stretch":
        canvas.create_oval(cx-5,cy-h,cx+5,cy-h+10, fill=color,outline="")
        canvas.create_line(cx,cy-h+10,cx,cy+4, fill=color,width=2)
        canvas.create_line(cx-h,cy-4,cx+h,cy-4, fill=color,width=2)
        canvas.create_line(cx,cy+4,cx-6,cy+h, fill=color,width=2)
        canvas.create_line(cx,cy+4,cx+6,cy+h, fill=color,width=2)

# ── Health Excel helpers ───────────────────────────────────────────────────────
def _load_health_log():
    path  = config.HEALTH_DATA
    today = str(date.today())
    counts = {"water": 0, "break": 0, "stretch": 0}
    if not os.path.exists(path):
        return counts
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if row and str(row[0]) == today and row[1] in counts:
                counts[str(row[1])] += 1
    except Exception:
        pass
    return counts

def _log_health_entry(entry_type: str):
    path = config.HEALTH_DATA
    os.makedirs(config.DATA_DIR, exist_ok=True)
    try:
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Health Log"
            ws.append(["date", "type", "time"])
        ws.append([str(date.today()), entry_type, datetime.now().strftime("%H:%M")])
        wb.save(path)
    except Exception as e:
        print(f"[health] log error: {e}")

# ── Cat notification popup (Phase 4) ─────────────────────────────────────────
class CatNotification:
    """
    A frameless popup with the cat image and a speech bubble message.
    Appears in the bottom-right corner, auto-dismisses after timeout.
    """
    POPUP_W = 320
    POPUP_H = 160

    def __init__(self, root: tk.Tk, spritesheet_path: str):
        self._root       = root
        self._sheet_path = spritesheet_path
        self._cat_photo  = None
        self._win        = None
        self._after_id   = None
        self._load_cat_image()

    def _load_cat_image(self):
        """Load sitting frame from active character spritesheet — always live."""
        try:
            from PIL import Image, ImageTk
            from features.character import get_active_character
            active = get_active_character()
            self._sheet_path = active
            sheet = Image.open(active).convert("RGBA")
            w, h  = sheet.size
            S     = config.SPRITE_SIZE
            if h >= S * 5 and w >= S * 4:
                frame = sheet.crop((0, 4*S, S, 5*S))
            else:
                sq = min(w, h)
                frame = sheet.crop(((w-sq)//2, (h-sq)//2, (w+sq)//2, (h+sq)//2))
            frame = frame.resize((72, 72), Image.LANCZOS)
            bg_hex = _theme_mod.get("CARD").lstrip("#")
            br,bg_,bb = int(bg_hex[0:2],16),int(bg_hex[2:4],16),int(bg_hex[4:6],16)
            bg = Image.new("RGBA", (72, 72), (br, bg_, bb, 255))
            bg.paste(frame, (0, 0), frame)
            self._cat_photo     = ImageTk.PhotoImage(bg)
            self._loaded_path   = active
            import theme as _lt
            self._loaded_theme  = _lt.name()   # track theme so bg stays correct
        except Exception as e:
            print(f"[notification] image load: {e}")
            self._cat_photo   = None
            self._loaded_path = None

    def show(self, title: str, message: str, color: str = ACCENT, timeout_ms: int = 6000):
        """Show a notification popup with the active character image."""
        # Always reload image — ensures theme bg colour is always current
        self._load_cat_image()
        # Cancel any existing popup
        if self._win and self._win.winfo_exists():
            self._win.destroy()
        if self._after_id:
            try: self._root.after_cancel(self._after_id)
            except: pass

        sw = self._root.winfo_screenwidth()
        sh = self._root.winfo_screenheight()
        wx = sw - self.POPUP_W - 20
        wy = sh - self.POPUP_H - 80   # above taskbar

        win = tk.Toplevel(self._root)
        win.overrideredirect(True)
        win.attributes("-topmost", True)
        win.geometry(f"{self.POPUP_W}x{self.POPUP_H}+{wx}+{wy}")
        win.config(highlightthickness=0, bd=0)
        import theme as _pt
        _BG=_pt.get("BG"); _CARD=_pt.get("CARD"); _BORDER=_pt.get("BORDER")
        _TEXT=_pt.get("TEXT"); _SUB=_pt.get("SUB"); _ACCENT=_pt.get("ACCENT")
        win.configure(bg=_CARD)
        # Suppress Windows DWM system border/shadow entirely
        try:
            import ctypes
            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                ctypes.windll.user32.GetParent(win.winfo_id()),
                2, ctypes.byref(ctypes.c_int(1)), ctypes.sizeof(ctypes.c_int))
        except Exception: pass
        self._win = win

        inner = tk.Frame(win, bg=_CARD, bd=0, highlightthickness=0)
        inner.place(x=0, y=0, width=self.POPUP_W, height=self.POPUP_H)

        # Character image on left
        if self._cat_photo:
            cat_lbl = tk.Label(inner, image=self._cat_photo, bg=_CARD)
            cat_lbl.place(x=10, y=(self.POPUP_H-72)//2)

        # Speech bubble area
        bubble_x = 92
        bubble_w = self.POPUP_W - bubble_x - 10

        # Bubble with 1px border
        bubble = tk.Frame(inner, bg=_BORDER, bd=0)
        bubble.place(x=bubble_x, y=10, width=bubble_w, height=self.POPUP_H-20)
        bubble_inner = tk.Frame(bubble, bg=_BG, bd=0)
        bubble_inner.place(x=1, y=1, width=bubble_w-2, height=self.POPUP_H-22)

        # Bubble tail (small triangle pointing left)
        tail = tk.Canvas(inner, width=10, height=14, bg=_CARD,
                         bd=0, highlightthickness=0)
        tail.place(x=bubble_x-10, y=24)
        tail.create_polygon(10,0, 10,14, 0,7, fill=_BORDER, outline="")
        tail2 = tk.Canvas(inner, width=8, height=12, bg=_CARD,
                          bd=0, highlightthickness=0)
        tail2.place(x=bubble_x-8, y=25)
        tail2.create_polygon(8,0, 8,12, 0,6, fill=_BG, outline="")

        f_title = tkfont.Font(family="Segoe UI", size=10, weight="bold")
        f_msg   = tkfont.Font(family="Segoe UI", size=9)
        f_btn   = tkfont.Font(family="Segoe UI", size=8)

        tk.Label(bubble_inner, text=title,
                 font=f_title, bg=_BG, fg=_ACCENT,
                 anchor="w").place(x=10, y=10, width=bubble_w-22)
        tk.Label(bubble_inner, text=message,
                 font=f_msg, bg=_BG, fg=_TEXT,
                 anchor="w", wraplength=bubble_w-22,
                 justify="left").place(x=10, y=32, width=bubble_w-22)

        def dismiss():
            if self._after_id:
                try: self._root.after_cancel(self._after_id)
                except: pass
            if win.winfo_exists():
                win.destroy()

        tk.Button(bubble_inner, text="OK  ✓",
                  font=f_btn, bg=_BORDER, fg=_TEXT,
                  activebackground=color, activeforeground=_BG,
                  relief="flat", cursor="hand2", bd=0,
                  command=dismiss).place(
                      x=bubble_w-60, y=self.POPUP_H-4-30-22,
                      width=50, height=20)

        # Click anywhere to dismiss
        for w in (win, inner, bubble_inner):
            w.bind("<Button-1>", lambda e: dismiss())

        self._after_id = self._root.after(timeout_ms, dismiss)


# ── Dashboard ─────────────────────────────────────────────────────────────────
class Dashboard:
    WIN_W = 500
    WIN_H = 660

    def __init__(self, root: tk.Tk, user: dict, on_close=None):
        self._root        = root
        self._user        = user
        self._on_close    = on_close
        self._win         = None
        self._content_frame = None

        # Health counts & labels
        self._health_counts = {"water": 0, "break": 0, "stretch": 0}
        self._hw_labels     = {}
        self._health_engine    = None   # set by App after init
        self._cat_widget_ref   = None   # set by App after init

        # Drag state for custom title bar
        self._drag_x = 0
        self._drag_y = 0

        # Cat notification popup
        self._notif = CatNotification(root, config.SPRITESHEET)

    # ── Public ─────────────────────────────────────────────────────────────────

    def set_health_reminder(self, engine):
        """Called by App to wire the standalone reminder engine."""
        self._health_engine = engine

    def _on_health_reminder_fired(self, key: str):
        """Called by HealthReminder (on main thread) when a reminder fires."""
        self._health_counts[key] = self._health_counts.get(key, 0) + 1
        self._refresh_count(key)

    def update_weather_city(self, city: str):
        """Called by settings when city is saved — updates weather tile subtitle live."""
        tile_lbl = getattr(self, "_weather_sub_lbl", None)
        if tile_lbl:
            try:
                if tile_lbl.winfo_exists():
                    tile_lbl.config(text=city)
            except Exception:
                pass

    def on_interval_change(self, key: str, minutes: int):
        """Delegate to the health engine and refresh the health tile label."""
        if self._health_engine:
            self._health_engine.set_interval(key, minutes)
        # Refresh the "next in X min" label if health view is open
        lbl = self._hw_labels.get(f"next_{key}")
        if lbl:
            try:
                if lbl.winfo_exists():
                    lbl.config(text=f"next in {minutes} min")
            except Exception:
                pass

    def open(self, cat_x: int = 100):
        if self._win and self._win.winfo_exists():
            try:
                if self._win.state() == "withdrawn":
                    self._win.deiconify()
                    self._win.lift()
                    if hasattr(self,"_restore_btn") and self._restore_btn:
                        try: self._restore_btn.destroy()
                        except: pass
                        self._restore_btn = None
                    return
            except Exception: pass
            self._win.lift(); return

        # Load today's health counts BEFORE building UI (fixes null values)
        self._health_counts = _load_health_log()

        self._win = tk.Toplevel(self._root)
        self._win.overrideredirect(True)          # ← removes ALL default chrome
        self._win.attributes("-topmost", True)
        self._win.configure(bg=BG)

        sw = self._root.winfo_screenwidth()
        sh = self._root.winfo_screenheight()
        wx = max(10, min(sw - self.WIN_W - 10, cat_x - self.WIN_W // 2))
        wy = max(10, sh - config.DISPLAY_SIZE - 60 - self.WIN_H - 10)
        self._win.geometry(f"{self.WIN_W}x{self.WIN_H}+{wx}+{wy}")

        self._build_shell()
        self._show_home()
        if self._health_engine:
            self._health_engine.register_ui(self._on_health_reminder_fired)

    def close(self):
        if self._health_engine:
            self._health_engine.unregister_ui()
        if self._win and self._win.winfo_exists():
            self._win.destroy()
            self._win = None

    def notify(self, title: str, message: str, color: str = ACCENT):
        """Show a cat speech-bubble notification popup."""
        self._notif.show(title, message, color)

    # ── Shell ──────────────────────────────────────────────────────────────────

    def _build_shell(self):
        _refresh_colours()
        w = self._win
        w.configure(bg=theme.get('BG'))

        # Rounded corners via Windows 11 DWM API
        try:
            import ctypes
            DWMWA_WINDOW_CORNER_PREFERENCE = 33
            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                ctypes.windll.user32.GetParent(w.winfo_id()),
                DWMWA_WINDOW_CORNER_PREFERENCE,
                ctypes.byref(ctypes.c_int(2)),
                ctypes.sizeof(ctypes.c_int)
            )
        except Exception:
            pass

        # ── Custom title bar ───────────────────────────────────────────────────
        titlebar = tk.Frame(w, bg=CARD, height=40)
        titlebar.pack(fill="x")
        titlebar.pack_propagate(False)

        # App icon dot
        dot = tk.Canvas(titlebar, width=12, height=12,
                        bg=CARD, bd=0, highlightthickness=0)
        dot.pack(side="left", padx=(12,6), pady=14)
        dot.create_oval(0,0,12,12, fill=ACCENT, outline="")

        # Title text
        f_title = tkfont.Font(family="Segoe UI", size=11, weight="bold")
        self._title_lbl = tk.Label(titlebar, text="Companion Dashboard",
                                   font=f_title, bg=CARD, fg=ACCENT)
        self._title_lbl.pack(side="left", pady=10)

        # Window controls — minimise, kill, close (right side)
        f_ctrl = tkfont.Font(family="Segoe UI", size=12)

        def _tooltip(widget, text):
            """Attach a simple hover tooltip to a widget."""
            tip = None
            def _show(e):
                nonlocal tip
                tip = tk.Toplevel(widget)
                tip.overrideredirect(True)
                tip.attributes("-topmost", True)
                tip.configure(bg=BORDER)
                tk.Label(tip, text=text,
                         font=tkfont.Font(family="Segoe UI", size=8),
                         bg=CARD, fg=TEXT, padx=6, pady=3).pack()
                tip.geometry(f"+{e.x_root+8}+{e.y_root+18}")
            def _hide(e):
                nonlocal tip
                if tip:
                    try: tip.destroy()
                    except: pass
                    tip = None
            widget.bind("<Enter>", _show)
            widget.bind("<Leave>", _hide)

        # Close button (✕)
        btn_close = tk.Button(titlebar, text="✕",
                  font=f_ctrl, bg=CARD, fg=SUB,
                  activebackground=DANGER, activeforeground="white",
                  relief="flat", cursor="hand2", bd=0, width=3,
                  command=self._handle_close)
        btn_close.pack(side="right", pady=8, padx=4)
        _tooltip(btn_close, "Close dashboard")

        # Kill button (⏻) — stops entire app
        btn_kill = tk.Button(titlebar, text="⏻",
                  font=f_ctrl, bg=CARD, fg="#E24B4A",
                  activebackground="#E24B4A", activeforeground="white",
                  relief="flat", cursor="hand2", bd=0, width=3,
                  command=self._kill_app)
        btn_kill.pack(side="right", pady=8)
        _tooltip(btn_kill, "Quit app completely")

        # Minimise button (─)
        btn_min = tk.Button(titlebar, text="─",
                  font=f_ctrl, bg=CARD, fg=SUB,
                  activebackground=BORDER, activeforeground=TEXT,
                  relief="flat", cursor="hand2", bd=0, width=3,
                  command=self._minimise)
        btn_min.pack(side="right", pady=8)
        _tooltip(btn_min, "Minimise dashboard")


        # Back button (hidden on home)
        self._back_btn = tk.Button(
            titlebar, text="←",
            font=tkfont.Font(family="Segoe UI", size=13),
            bg=CARD, fg=SUB,
            activebackground=BORDER, activeforeground=TEXT,
            relief="flat", cursor="hand2", bd=0, width=2,
            command=self._show_home,
        )
        _tooltip(self._back_btn, "Back to home")

        # Drag the window via title bar
        titlebar.bind("<ButtonPress-1>",  self._drag_start)
        titlebar.bind("<B1-Motion>",       self._drag_motion)
        self._title_lbl.bind("<ButtonPress-1>",  self._drag_start)
        self._title_lbl.bind("<B1-Motion>",       self._drag_motion)
        dot.bind("<ButtonPress-1>",  self._drag_start)
        dot.bind("<B1-Motion>",       self._drag_motion)

        # Separator
        tk.Frame(w, bg=BORDER, height=1).pack(fill="x")

        # Swappable content area
        self._content_area = tk.Frame(w, bg=BG)
        self._content_area.pack(fill="both", expand=True)

    def _drag_start(self, event):
        self._drag_x = event.x_root - self._win.winfo_x()
        self._drag_y = event.y_root - self._win.winfo_y()

    def _drag_motion(self, event):
        x = event.x_root - self._drag_x
        y = event.y_root - self._drag_y
        self._win.geometry(f"+{x}+{y}")

    def _minimise(self):
        """Hide dashboard; show a small restore button near the taskbar."""
        win = self._win
        if not win or not win.winfo_exists(): return
        win.withdraw()
        if self._cat_widget_ref and hasattr(self._cat_widget_ref, 'resume'):
            self._cat_widget_ref.resume()
        self._show_restore_button()


    def _show_restore_button(self):
        """Small floating restore button shown while dashboard is minimised."""
        if hasattr(self, '_restore_btn') and self._restore_btn and                 self._restore_btn.winfo_exists():
            return
        root = self._root
        sw = root.winfo_screenwidth()
        sh = root.winfo_screenheight()
        btn = tk.Toplevel(root)
        btn.overrideredirect(True)
        btn.attributes("-topmost", True)
        btn.geometry(f"44x28+{sw-50}+{sh-80}")
        btn.configure(bg=CARD)
        self._restore_btn = btn

        def restore():
            if btn.winfo_exists(): btn.destroy()
            if self._win and self._win.winfo_exists():
                self._win.deiconify()
                self._win.lift()
            # Stop cat when dashboard is opened via restore icon
            if self._cat_widget_ref:
                try:
                    self._cat_widget_ref._stopped = True
                    self._cat_widget_ref._set_state("stopped")
                except Exception: pass

        tk.Button(btn, text="🐱", font=tkfont.Font(family="Segoe UI", size=12),
                  bg=CARD, fg=ACCENT, relief="flat", cursor="hand2", bd=0,
                  command=restore).pack(fill="both", expand=True)

    def _swap_content(self, builder_fn, title="", show_back=True):
        for child in self._content_area.winfo_children():
            child.destroy()

        self._title_lbl.config(text=title or "Companion Dashboard")

        if show_back:
            self._back_btn.pack(side="left", padx=(4,0), pady=8)
        else:
            self._back_btn.pack_forget()

        self._content_frame = tk.Frame(self._content_area, bg=BG)
        self._content_frame.pack(fill="both", expand=True, padx=16, pady=14)
        builder_fn(self._content_frame)

    # ── Home ──────────────────────────────────────────────────────────────────

    def update_tile_subtitle(self, key: str, sub: str):
        """Update a tile's subtitle — updates TILES dict and the live label."""
        for i, tile in enumerate(TILES):
            if tile["key"] == key:
                TILES[i]["sub"] = sub
                break
        # Update the live weather subtitle label immediately (before home rebuild)
        if key == "weather":
            lbl = getattr(self, "_weather_sub_lbl", None)
            if lbl:
                try:
                    if lbl.winfo_exists():
                        lbl.config(text=sub)
                except Exception:
                    pass
            # Also force-refresh the home grid if we are on home right now
            # so the tile shows updated city without needing to navigate away
            try:
                if hasattr(self, "_content_frame") and self._content_frame.winfo_exists():
                    # Check if home is currently shown (back button is hidden)
                    if not self._back_btn.winfo_ismapped():
                        self._show_home()
            except Exception:
                pass

    def _show_home(self):
        self._swap_content(self._build_home, title="Companion Dashboard", show_back=False)

    @staticmethod
    def _get_weather_city_display_static():
        """Read saved city from user.xlsx for tile display. Falls back to Supabase."""
        try:
            import openpyxl as _xl
            if os.path.exists(config.USER_DATA):
                _wb = _xl.load_workbook(config.USER_DATA)
                _ws = _wb.active
                for _row in _ws.iter_rows(values_only=True):
                    if _row and _row[0] == "weather_city" and _row[1]:
                        city = str(_row[1]).strip()
                        return city if ", " in city else f"{city}, IN"
        except Exception as e:
            print(f"[dashboard] city read error: {e}")
        try:
            import supabase_config
            city = supabase_config.get_key("City", "Bengaluru")
            return f"{city}, IN" if city else "Bengaluru, IN"
        except Exception:
            return "Bengaluru, IN"

    def _build_home(self, parent):
        _refresh_colours()
        name    = self._user.get("name", "Friend")
        day_str = datetime.now().strftime("%A, %d %B")

        # Greeting
        banner_border = tk.Frame(parent, bg=BORDER)
        banner_border.pack(fill="x", pady=(0, 14))
        banner = tk.Frame(banner_border, bg=CARD)
        banner.pack(fill="x", padx=1, pady=1)

        char_icon = theme.icon()
        tk.Label(banner, text=f"Hello, {name}!  {char_icon}",
                 font=tkfont.Font(family="Segoe UI", size=20, weight="bold"),
                 bg=CARD, fg=ACCENT, anchor="w").pack(fill="x", padx=16, pady=(12,2))
        tk.Label(banner, text=f"{day_str}  ·  Your cat is keeping you company",
                 font=tkfont.Font(family="Segoe UI", size=10),
                 bg=CARD, fg=SUB, anchor="w").pack(fill="x", padx=16, pady=(0,12))

        # Tile grid — 3 cols, auto rows, no expand so health bar stays visible
        grid = tk.Frame(parent, bg=BG)
        grid.pack(fill="x")
        COLS = 3
        for i, tile in enumerate(TILES):
            self._make_tile(grid, tile, i//COLS, i%COLS, 138, 118, 8)
        for c in range(COLS):
            grid.columnconfigure(c, weight=1)

        # Health bar — uses loaded counts (not null)
        hbar_border = tk.Frame(parent, bg=BORDER)
        hbar_border.pack(fill="x", pady=(12, 0))
        hbar = tk.Frame(hbar_border, bg=CARD)
        hbar.pack(fill="x", padx=1, pady=1)

        f_h = tkfont.Font(family="Segoe UI", size=9)
        for idx, (key, icon, label) in enumerate(
                [("water","💧","Water"),("break","☕","Break"),("stretch","🧘","Stretch")]):
            cell = tk.Frame(hbar, bg=CARD)
            cell.pack(side="left", expand=True, fill="x")
            tk.Label(cell, text=f"{icon} {label}:",
                     font=f_h, bg=CARD, fg=SUB).pack(side="left", padx=(12,4), pady=8)
            cnt = self._health_counts.get(key, 0)
            val = f"{cnt}x today" if cnt else "—"
            lbl = tk.Label(cell, text=val, font=f_h, bg=CARD, fg=ACCENT)
            lbl.pack(side="left")
            setattr(self, f"_hbar_{key}", lbl)

    def _make_tile(self, parent, tile, row, col, tw, th, gap):
        color = tile["color"]
        border = tk.Frame(parent, bg=BORDER, cursor="hand2")
        border.grid(row=row, column=col, padx=gap//2, pady=gap//2, sticky="nsew")
        card = tk.Frame(border, bg=CARD, cursor="hand2",
                        width=tw-2, height=th-2)
        card.pack(fill="both", expand=True, padx=1, pady=1)
        card.pack_propagate(False)

        ic = tk.Canvas(card, width=36, height=36, bg=CARD,
                       bd=0, highlightthickness=0, cursor="hand2")
        ic.place(x=12, y=12)
        draw_icon(ic, tile["key"], 18, 18, 26, color)

        lbl = tk.Label(card, text=tile["label"],
                       font=tkfont.Font(family="Segoe UI", size=11, weight="bold"),
                       bg=CARD, fg=TEXT, anchor="w", cursor="hand2")
        lbl.place(x=12, y=56)
        # Read live subtitle — for weather tile, always read from user.xlsx
        sub_text = tile["sub"]
        if tile.get("sub_key") == "weather_city":
            sub_text = Dashboard._get_weather_city_display_static()

        sub = tk.Label(card, text=sub_text,
                       font=tkfont.Font(family="Segoe UI", size=9),
                       bg=CARD, fg=SUB, anchor="w", cursor="hand2")
        sub.place(x=12, y=78)

        # Store ref for live update (weather city)
        if tile.get("sub_key") == "weather_city":
            self._weather_sub_lbl = sub

        all_w = [border, card, ic, lbl, sub]

        def on_enter(e=None, b=border, c=card, i=ic, l=lbl, s=sub, col=color):
            _refresh_colours()
            b.config(bg=col); c.config(bg=col); i.config(bg=col)
            i.delete("all"); draw_icon(i, tile["key"], 18, 18, 26, BG)
            l.config(bg=col, fg=BG); s.config(bg=col, fg=BG)

        def on_leave(e=None, b=border, c=card, i=ic, l=lbl, s=sub, col=color):
            _refresh_colours()
            b.config(bg=BORDER); c.config(bg=CARD); i.config(bg=CARD)
            i.delete("all"); draw_icon(i, tile["key"], 18, 18, 26, col)
            l.config(bg=CARD, fg=TEXT); s.config(bg=CARD, fg=SUB)

        for wgt in all_w:
            wgt.bind("<Enter>",    on_enter)
            wgt.bind("<Leave>",    on_leave)
            wgt.bind("<Button-1>", lambda e=None, k=tile["key"]: self._route(k))

    # ── Routing ───────────────────────────────────────────────────────────────

    def _route(self, key: str):
        {
            "health":   self._show_health,
            "planner":  self._show_planner,
            "notes":    self._show_notes,
            "weather":  self._show_weather,
            "pomodoro": self._show_pomodoro,
            "todo":     self._show_todo,
            "ai_chat":  self._show_ai_chat,
            "settings":   self._show_settings,
            "character":  self._show_character,
        }.get(key, lambda: None)()

    def _show_todo(self):
        from features import todo as todo_mod
        def builder(parent):
            todo_mod.build(parent, self._root)
        self._swap_content(builder, title="To-Do List")

    def _show_notes(self):
        from features import notes as notes_mod
        def builder(parent):
            notes_mod.build(parent, self._root)
        self._swap_content(builder, title="Notes")

    def _show_planner(self):
        from features import planner as planner_mod
        def builder(parent):
            planner_mod.build(parent, self._root, notify_fn=self.notify)
        self._swap_content(builder, title="Daily Planner")

    # ── Health view ───────────────────────────────────────────────────────────

    def _show_health(self):
        # Refresh counts from Excel before showing
        self._health_counts = _load_health_log()
        self._swap_content(self._build_health, title="Health Tracker")

    def _build_health(self, parent):
        self._hw_labels = {}
        f_head = tkfont.Font(family="Segoe UI", size=12, weight="bold")
        f_cnt  = tkfont.Font(family="Segoe UI", size=22, weight="bold")
        f_btn  = tkfont.Font(family="Segoe UI", size=10, weight="bold")
        f_sub  = tkfont.Font(family="Segoe UI", size=9)
        f_next = tkfont.Font(family="Segoe UI", size=9)

        reminders = [
            ("water",   "water",   "Drink Water",   config.WATER_INTERVAL_MIN,   "#4FC3F7",
             "Stay hydrated — drink a glass of water."),
            ("break",   "break_",  "Short Break",   config.BREAK_INTERVAL_MIN,   "#FFB74D",
             "Rest your eyes and take a short break."),
            ("stretch", "stretch", "Stretch",        config.STRETCH_INTERVAL_MIN, "#81C784",
             "Stand up and stretch for 2 minutes."),
        ]

        for key, icon_key, label, interval, color, tip in reminders:
            card_border = tk.Frame(parent, bg=BORDER)
            card_border.pack(fill="x", pady=6)
            card = tk.Frame(card_border, bg=CARD)
            card.pack(fill="x", padx=1, pady=1)

            row = tk.Frame(card, bg=CARD)
            row.pack(fill="x", padx=14, pady=12)

            left = tk.Frame(row, bg=CARD)
            left.pack(side="left", fill="y")

            ic = tk.Canvas(left, width=40, height=40, bg=CARD,
                           bd=0, highlightthickness=0)
            ic.pack()
            draw_icon(ic, icon_key, 20, 20, 28, color)

            tk.Label(left, text=label, font=f_head, bg=CARD, fg=TEXT).pack(anchor="w")
            tk.Label(left, text=tip, font=f_sub, bg=CARD, fg=SUB,
                     wraplength=220, justify="left").pack(anchor="w")

            right = tk.Frame(row, bg=CARD)
            right.pack(side="right", anchor="ne")

            # Use loaded counts — never null
            cnt = self._health_counts.get(key, 0)
            cnt_lbl = tk.Label(right, text=str(cnt), font=f_cnt, bg=CARD, fg=color)
            cnt_lbl.pack(anchor="e")
            tk.Label(right, text="today", font=f_sub, bg=CARD, fg=SUB).pack(anchor="e")

            tk.Button(right, text="Log  +1",
                      font=f_btn, bg=color, fg="#1A1B2E",
                      activebackground=BORDER, activeforeground=TEXT,
                      relief="flat", cursor="hand2", bd=0,
                      command=lambda k=key, l=cnt_lbl, c=color: self._log_manual(k,l,c)
                      ).pack(anchor="e", pady=(6,0), ipadx=8, ipady=4)

            # Read from engine if available, else show interval default
            if self._health_engine:
                secs_left = self._health_engine._countdown.get(key, interval*60)
                mins = max(1, secs_left // 60)
            else:
                mins = interval
            next_lbl = tk.Label(right, text=f"next in {mins} min",
                                font=f_next, bg=CARD, fg=SUB)
            next_lbl.pack(anchor="e", pady=(4,0))
            self._hw_labels[f"next_{key}"] = next_lbl
            self._hw_labels[f"cnt_{key}"]  = cnt_lbl

        # Summary
        tk.Frame(parent, bg=BORDER, height=1).pack(fill="x", pady=(10,6))
        summary_row = tk.Frame(parent, bg=BG)
        summary_row.pack(fill="x")
        tk.Label(summary_row, text="Today's summary",
                 font=f_head, bg=BG, fg=SUB).pack(side="left")
        for key, _, label, *_ in reminders:
            cnt = self._health_counts.get(key, 0)
            chip = tk.Frame(summary_row, bg=CARD)
            chip.pack(side="right", padx=4)
            tk.Label(chip, text=f"{cnt}x {label}",
                     font=tkfont.Font(family="Segoe UI", size=9),
                     bg=CARD, fg=TEXT, padx=10, pady=4).pack()

    def _log_manual(self, key: str, cnt_lbl: tk.Label, color: str):
        _log_health_entry(key)
        self._health_counts[key] = self._health_counts.get(key, 0) + 1
        cnt = self._health_counts[key]
        cnt_lbl.config(text=str(cnt))
        bar_lbl = getattr(self, f"_hbar_{key}", None)
        if bar_lbl:
            try:
                if bar_lbl.winfo_exists():
                    bar_lbl.config(text=f"{cnt}x today")
            except: pass


    def _refresh_count(self, key: str):
        cnt = self._health_counts.get(key, 0)
        lbl = self._hw_labels.get(f"cnt_{key}")
        if lbl:
            try:
                if lbl.winfo_exists(): lbl.config(text=str(cnt))
            except: pass
        bar = getattr(self, f"_hbar_{key}", None)
        if bar:
            try:
                if bar.winfo_exists(): bar.config(text=f"{cnt}x today")
            except: pass

    # ── Placeholder ───────────────────────────────────────────────────────────

    def _show_pomodoro(self):
        from features import pomodoro as feat_pomodoro
        def builder(p):
            feat_pomodoro.build(p, self._root, notify_fn=self.notify)
        self._swap_content(builder, title="Pomodoro Timer")

    def _show_weather(self):
        from features import weather as feat_weather
        self._swap_content(
            lambda p: feat_weather.build(p, self._root),
            title="Weather")

    def _show_ai_chat(self):
        from features import ai_chat as feat_ai
        def builder(p):
            feat_ai.build(p, self._root, user=self._user)
        self._swap_content(builder, title="AI Chat  🐱")

    def _show_settings(self):
        from features import settings as feat_settings
        def builder(p):
            feat_settings.build(
                p, self._root,
                user=self._user,
                on_name_change=self._on_name_change,
                on_interval_change=self.on_interval_change,
                on_tile_update=self.update_tile_subtitle,
            )
        self._swap_content(builder, title="Settings")

    def _on_name_change(self, new_name: str):
        """Called by settings when user updates their name."""
        self._user["name"] = new_name

    def _show_character(self):
        from features import character as feat_char
        def builder(p):
            feat_char.build(
                p, self._root,
                on_character_change=self._on_character_change,
            )
        self._swap_content(builder, title="Choose Companion")

    def _on_character_change(self, path: str):
        """Hot-swap character + apply matching theme + rebuild dashboard."""
        # 1. Apply theme using is_predefined() — works for ALL named themes
        if theme.is_predefined(path):
            # cat / dragon / forest / sakura / midnight → auto-apply mapped theme
            theme.apply(theme.theme_for_character(path))
            self._save_user_theme_override("")
        else:
            # Custom/unknown character → use saved override or "default"
            import openpyxl as _oxl, config as _cfg2, os as _dos
            override = ""
            try:
                if _dos.path.exists(_cfg2.USER_DATA):
                    _wb = _oxl.load_workbook(_cfg2.USER_DATA)
                    for r in _wb.active.iter_rows(values_only=True):
                        if r[0] == "theme_override" and r[1]:
                            override = str(r[1]); break
            except Exception: pass
            theme.apply(override if override in theme.THEMES else "default")
        # 2. Hot-swap sprite
        if self._cat_widget_ref and hasattr(self._cat_widget_ref, 'reload_character'):
            self._cat_widget_ref.reload_character(path)
        # 3. Rebuild dashboard safely via after() to avoid TclError
        if self._win and self._win.winfo_exists():
            self._root.after(50, self._rebuild_with_theme)

    def _save_user_theme_override(self, value: str):
        """Clear or set theme_override in user.xlsx."""
        import openpyxl, config as _cfg, os as _os
        _os.makedirs(_cfg.DATA_DIR, exist_ok=True)
        try:
            if _os.path.exists(_cfg.USER_DATA):
                wb = openpyxl.load_workbook(_cfg.USER_DATA)
                ws = wb.active
                for row in ws.iter_rows():
                    if row[0].value == "theme_override":
                        row[1].value = value; wb.save(_cfg.USER_DATA); return
                ws.append(["theme_override", value]); wb.save(_cfg.USER_DATA)
        except Exception as e:
            print(f"[dashboard] save theme_override: {e}")

    def _rebuild_with_theme(self):
        """Destroy and recreate the entire dashboard window with new theme.
        This is the only reliable way to retheme every pixel including
        the titlebar, separator, buttons, and all content frames.
        """
        if not (self._win and self._win.winfo_exists()):
            return

        # Remember position so window reopens in same place
        try:
            wx = self._win.winfo_x()
            wy = self._win.winfo_y()
        except Exception:
            wx, wy = None, None

        # Destroy ALL widgets in the window completely
        for child in list(self._win.winfo_children()):
            try: child.destroy()
            except Exception: pass

        # Refresh colours then rebuild
        _refresh_colours()
        self._build_shell()

        # Restore position
        if wx is not None:
            try: self._win.geometry(f"{self.WIN_W}x{self.WIN_H}+{wx}+{wy}")
            except Exception: pass

        # Show home with new theme
        self._show_home()

    def set_cat_widget(self, cat_widget):
        """Called by App so dashboard can call reload_character."""
        self._cat_widget_ref = cat_widget

    def _show_placeholder(self, title: str, phase: str):
        def builder(parent):
            tk.Label(parent, text=title,
                     font=tkfont.Font(family="Segoe UI", size=16, weight="bold"),
                     bg=BG, fg=ACCENT).pack(pady=(50,8))
            tk.Label(parent, text=f"Coming in {phase}",
                     font=tkfont.Font(family="Segoe UI", size=11),
                     bg=BG, fg=SUB).pack()
        self._swap_content(builder, title=title)

    # ── Close / Kill ──────────────────────────────────────────────────────────

    def _handle_close(self):
        self.close()
        if self._on_close:
            self._on_close()

    def _kill_app(self):
        """Kill the entire app — stops health reminders and destroys root."""
        try:
            self._root.quit()
            self._root.destroy()
        except Exception:
            import sys
            sys.exit(0)
